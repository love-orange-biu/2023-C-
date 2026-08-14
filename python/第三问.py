from __future__ import annotations

import warnings
from pathlib import Path
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import font_manager
from matplotlib.lines import Line2D
from scipy.optimize import Bounds, LinearConstraint, milp
from scipy.sparse import coo_matrix


# =====================================================================
# 1. 路径和可调参数
# =====================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
OUTPUT_DIR = SCRIPT_DIR / "第三问"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

MODEL_FILE = SCRIPT_DIR / "C题_正确处理后建模数据.xlsx"
STEP5_FILE = (
    SCRIPT_DIR
    / "问题二_Step5_最终版"
    / "问题二_Step5_稳健补货定价优化结果.xlsx"
)
ELASTICITY_FILE = SCRIPT_DIR / "问题二_Step2" / "六大品类价格弹性估计.xlsx"
LOSS_FILE = SCRIPT_DIR / "附件4.xlsx"
FONT_FILE = PROJECT_DIR / "fonts" / "msyh.ttc"

DECISION_DATE = pd.Timestamp("2023-07-01")
RECENT_START = pd.Timestamp("2023-06-24")
RECENT_END = pd.Timestamp("2023-06-30")
HISTORY_START = pd.Timestamp("2023-06-01")

MIN_ITEMS = 27
MAX_ITEMS = 33
MIN_DISPLAY = 2.5  # kg
SAFETY_RATE = 0.05
PRICE_GRID_SIZE = 7
PROFIT_TOLERANCE_YUAN = 0.01
SIGNIFICANT_PRICE_BAND = 0.12
INSIGNIFICANT_PRICE_BAND = 0.05
COST_DECAY_RATE = 0.25

CATEGORIES = ["花叶类", "花菜类", "水生根茎类", "茄类", "辣椒类", "食用菌"]
CATEGORY_COLORS = {
    "花叶类": "#7FA9C9",
    "花菜类": "#DDA66D",
    "水生根茎类": "#8DB79E",
    "茄类": "#D89591",
    "辣椒类": "#A99BC5",
    "食用菌": "#B29A80",
}


def configure_chinese_font() -> str:
    """显式注册项目自带的微软雅黑，避免换机后中文丢失。"""
    if not FONT_FILE.exists():
        raise FileNotFoundError(f"未找到微软雅黑字体：{FONT_FILE}")
    font_manager.fontManager.addfont(str(FONT_FILE))
    font_name = font_manager.FontProperties(fname=str(FONT_FILE)).get_name()
    plt.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": [font_name],
            "axes.unicode_minus": False,
            "figure.dpi": 120,
            "savefig.dpi": 320,
            "axes.edgecolor": "#666666",
            "axes.labelcolor": "#444444",
            "xtick.color": "#555555",
            "ytick.color": "#555555",
        }
    )
    return font_name


def require_inputs() -> None:
    for path in [MODEL_FILE, STEP5_FILE, ELASTICITY_FILE, LOSS_FILE, FONT_FILE]:
        if not path.exists():
            raise FileNotFoundError(f"缺少输入文件：{path}")


def read_inputs() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    candidates = pd.read_excel(MODEL_FILE, sheet_name="问题3候选单品")
    daily = pd.read_excel(MODEL_FILE, sheet_name="单品日数据")
    wholesale = pd.read_excel(MODEL_FILE, sheet_name="单品日批发价")
    category_plan = pd.read_excel(STEP5_FILE, sheet_name="每日最优策略")
    elasticity = pd.read_excel(ELASTICITY_FILE)
    item_loss = pd.read_excel(LOSS_FILE, sheet_name="Sheet1")

    for frame in [candidates, daily, wholesale, item_loss]:
        frame["单品编码"] = frame["单品编码"].astype("int64").astype(str)
    daily["日期"] = pd.to_datetime(daily["日期"])
    wholesale["日期"] = pd.to_datetime(wholesale["日期"])
    category_plan["日期"] = pd.to_datetime(category_plan["日期"])
    item_loss["损耗率"] = pd.to_numeric(item_loss["损耗率(%)"], errors="coerce") / 100

    category_plan = category_plan[
        category_plan["日期"].eq(DECISION_DATE)
        & category_plan["分类名称"].isin(CATEGORIES)
    ].copy()
    if len(category_plan) != len(CATEGORIES):
        raise ValueError("Step5 中 2023-07-01 的六大品类策略不完整。")
    if len(candidates) != 49:
        warnings.warn(f"候选单品数为 {len(candidates)}，与预期的 49 不同。")

    return candidates, daily, wholesale, category_plan, item_loss, elasticity


def weighted_recent_cost(group: pd.DataFrame) -> float:
    """对最近批发价做指数加权，越接近 6 月 30 日权重越高。"""
    temp = group.sort_values("日期").dropna(subset=["批发价格(元/千克)"])
    if temp.empty:
        return np.nan
    days_ago = (RECENT_END - temp["日期"]).dt.days.clip(lower=0)
    weights = np.exp(-COST_DECAY_RATE * days_ago.to_numpy())
    return float(np.average(temp["批发价格(元/千克)"], weights=weights))


def build_item_parameters(
    candidates: pd.DataFrame,
    daily: pd.DataFrame,
    wholesale: pd.DataFrame,
    category_plan: pd.DataFrame,
    item_loss: pd.DataFrame,
    elasticity: pd.DataFrame,
) -> pd.DataFrame:
    """
    将第二问品类需求按近 7 日销量份额分解到单品，并估计价格、成本和容量参数。
    """
    candidate_codes = set(candidates["单品编码"])
    d = daily[daily["单品编码"].isin(candidate_codes)].copy()
    for col in ["净销量(千克)", "销售额(元)", "加权平均售价(元/千克)"]:
        d[col] = pd.to_numeric(d[col], errors="coerce")

    recent = d[d["日期"].between(RECENT_START, RECENT_END)].copy()
    recent["正销量"] = recent["净销量(千克)"].clip(lower=0)
    recent["正销售额"] = np.where(recent["正销量"] > 0, recent["销售额(元)"].clip(lower=0), 0)

    basic = (
        recent.groupby(["单品编码", "单品名称", "分类名称"], as_index=False)
        .agg(
            近七日销量=("正销量", "sum"),
            近七日销售额=("正销售额", "sum"),
            近七日最大日销量=("正销量", "max"),
            有销量天数=("正销量", lambda s: int((s > 0).sum())),
        )
    )
    basic["参考售价(元/千克)"] = (
        basic["近七日销售额"] / basic["近七日销量"].replace(0, np.nan)
    )

    hist = d[d["日期"].between(HISTORY_START, RECENT_END)].copy()
    price_stats = (
        hist.groupby("单品编码")["加权平均售价(元/千克)"]
        .agg(
            近期售价标准差="std",
            近期售价10分位=lambda s: s.quantile(0.10),
            近期售价90分位=lambda s: s.quantile(0.90),
        )
        .reset_index()
    )
    basic = basic.merge(price_stats, on="单品编码", how="left")

    w = wholesale[
        wholesale["单品编码"].isin(candidate_codes)
        & wholesale["日期"].between(pd.Timestamp("2023-06-17"), RECENT_END)
    ].copy()
    raw_cost = w.groupby("单品编码").apply(weighted_recent_cost, include_groups=False)
    raw_cost = raw_cost.rename("近期批发价加权值").reset_index()
    basic = basic.merge(raw_cost, on="单品编码", how="left")
    basic = basic.merge(
        candidates[["单品编码", "最近批发价(元/千克)", "损耗率"]],
        on="单品编码",
        how="left",
        suffixes=("", "_品类备用"),
    )
    basic["近期批发价加权值"] = basic["近期批发价加权值"].fillna(
        basic["最近批发价(元/千克)"]
    )

    loss_map = item_loss.set_index("单品编码")["损耗率"]
    basic["单品损耗率"] = basic["单品编码"].map(loss_map)
    basic["单品损耗率"] = basic["单品损耗率"].fillna(basic["损耗率"])
    basic["单品损耗率"] = basic["单品损耗率"].clip(0, 0.50)

    plan_cols = [
        "分类名称",
        "价格弹性β",
        "预测基准销量(千克)",
        "预测批发价(元/千克)",
    ]
    basic = basic.merge(category_plan[plan_cols], on="分类名称", how="left")
    basic = basic.rename(columns={"预测基准销量(千克)": "品类目标需求(千克)"})
    basic = basic.merge(
        elasticity[["分类名称", "P值", "95%CI下限", "95%CI上限"]],
        on="分类名称",
        how="left",
    )

    basic["品类近七日销量"] = basic.groupby("分类名称")["近七日销量"].transform("sum")
    basic["近七日品类内份额"] = basic["近七日销量"] / basic["品类近七日销量"]
    basic["单品基准需求(千克)"] = basic["品类目标需求(千克)"] * basic["近七日品类内份额"]
    # 使单品批发价的份额加权平均与第二问品类预测一致。
    basic["原始成本份额乘积"] = basic["近期批发价加权值"] * basic["近七日品类内份额"]
    category_raw_cost = basic.groupby("分类名称")["原始成本份额乘积"].transform("sum")
    basic["品类成本校正系数"] = basic["预测批发价(元/千克)"] / category_raw_cost
    basic["预测单品批发价(元/千克)"] = (
        basic["近期批发价加权值"] * basic["品类成本校正系数"]
    )

    for idx, row in basic.iterrows():
        pref = float(row["参考售价(元/千克)"])
        basic.loc[idx, "近期售价标准差"] = max(
            float(row["近期售价标准差"]) if pd.notna(row["近期售价标准差"]) else 0,
            0.05 * pref,
        )
        if pd.isna(row["近期售价10分位"]):
            basic.loc[idx, "近期售价10分位"] = 0.90 * pref
        if pd.isna(row["近期售价90分位"]):
            basic.loc[idx, "近期售价90分位"] = 1.10 * pref

    basic["分类名称"] = pd.Categorical(basic["分类名称"], CATEGORIES, ordered=True)
    basic = basic.sort_values(["分类名称", "近七日销量"], ascending=[True, False]).reset_index(drop=True)
    basic["单品序号"] = np.arange(len(basic))
    return basic


def make_price_tiers(
    items: pd.DataFrame,
    significant_band: float = SIGNIFICANT_PRICE_BAND,
    insignificant_band: float = INSIGNIFICANT_PRICE_BAND,
) -> pd.DataFrame:
    rows: list[dict] = []
    for i, row in items.iterrows():
        pref = float(row["参考售价(元/千克)"])
        cost = float(row["预测单品批发价(元/千克)"])
        p10 = float(row["近期售价10分位"])
        p90 = float(row["近期售价90分位"])
        loss = float(row["单品损耗率"])
        significant = bool(float(row["P值"]) < 0.05)
        band = significant_band if significant else insignificant_band
        low = max((1 - band) * pref, min(p10, pref))
        high = min((1 + band) * pref, max(pref, p90))
        if high <= low:
            high = max(low + 0.20, 1.03 * low)

        prices = np.linspace(low, high, PRICE_GRID_SIZE)
        prices = np.r_[prices, np.clip(pref, low, high)]
        prices = np.unique(np.round(prices, 2))
        beta = float(row["价格弹性β"])
        sigma = float(row["近期售价标准差"])
        for price in prices:
            # 需求基准量由品类预测按近期销量份额分解得到。
            # 历史最大销量仅作参数诊断，不再把“容量”当作实际需求。
            demand = float(row["单品基准需求(千克)"]) * (price / pref) ** beta
            replenish = max(MIN_DISPLAY, (1 + SAFETY_RATE) * demand / max(1 - loss, 1e-6))
            revenue = float(price) * demand
            purchase_cost = cost * replenish
            profit = revenue - purchase_cost
            price_deviation_index = ((price - pref) / sigma) ** 2
            rows.append(
                {
                    "单品序号": i,
                    "单品编码": row["单品编码"],
                    "价格档位": float(price),
                    "价格决定的预测销量": max(demand, 0),
                    "档位建议补货量": replenish,
                    "档位预测销售额": revenue,
                    "档位预测采购成本": purchase_cost,
                    "档位预测经营利润": profit,
                    "标准化价格偏离平方": max(float(price_deviation_index), 0),
                    "价格下界": low,
                    "价格上界": high,
                    "弹性是否显著": "是" if significant else "否",
                }
            )
    tiers = pd.DataFrame(rows)
    tiers["档位序号"] = np.arange(len(tiers))
    return tiers


def solve_three_stage(items: pd.DataFrame, tiers: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame, dict]:
    """
    三阶段 MILP：
    1. 最大化最低品类满足率；
    2. 在最优服务水平下最大化真实经营利润；
    3. 在利润距最优值不超过 0.01 元时最小化价格偏离。
    """
    n_items = len(items)
    n_tiers = len(tiers)
    n_cats = len(CATEGORIES)
    theta_idx = n_tiers
    n_vars = n_tiers + 1

    item_of_tier = tiers["单品序号"].astype(int).to_numpy()
    tier_demand = tiers["价格决定的预测销量"].to_numpy(float)
    targets = np.array(
        [float(items.loc[items["分类名称"] == cat, "品类目标需求(千克)"].iloc[0]) for cat in CATEGORIES]
    )

    lower = np.zeros(n_vars)
    upper = np.ones(n_vars)
    integrality = np.zeros(n_vars, dtype=int)
    integrality[:n_tiers] = 1

    row_indices: list[int] = []
    col_indices: list[int] = []
    values: list[float] = []
    b_l: list[float] = []
    b_u: list[float] = []

    def add_constraint(coeffs: dict[int, float], low: float, high: float) -> None:
        r = len(b_l)
        for c, v in coeffs.items():
            if abs(v) > 0:
                row_indices.append(r)
                col_indices.append(c)
                values.append(float(v))
        b_l.append(low)
        b_u.append(high)

    add_constraint({j: 1 for j in range(n_tiers)}, MIN_ITEMS, MAX_ITEMS)

    for i in range(n_items):
        js = np.flatnonzero(item_of_tier == i)
        add_constraint({int(j): 1 for j in js}, 0, 1)

    for k, cat in enumerate(CATEGORIES):
        item_ids = items.index[items["分类名称"] == cat].to_numpy()
        js = np.flatnonzero(np.isin(item_of_tier, item_ids))
        # 价格档位一旦选中，其销量由需求函数唯一确定，不再是自由决策变量。
        served = {int(j): tier_demand[int(j)] for j in js}
        served[theta_idx] = -targets[k]
        add_constraint(served, 0, np.inf)
        # 品类预测销量不得超过第二问给出的市场需求。
        add_constraint({int(j): tier_demand[int(j)] for j in js}, -np.inf, targets[k])

    def linear_constraint() -> LinearConstraint:
        matrix = coo_matrix((values, (row_indices, col_indices)), shape=(len(b_l), n_vars)).tocsr()
        return LinearConstraint(matrix, np.asarray(b_l), np.asarray(b_u))

    stage1_c = np.zeros(n_vars)
    stage1_c[theta_idx] = -1
    result1 = milp(
        c=stage1_c,
        integrality=integrality,
        bounds=Bounds(lower, upper),
        constraints=linear_constraint(),
        options={"time_limit": 180, "mip_rel_gap": 1e-6},
    )
    if not result1.success:
        raise RuntimeError(f"第一阶段求解失败：{result1.message}")
    best_theta = float(result1.x[theta_idx])

    # 第二阶段：在最优服务水平下直接最大化真实经营利润。
    add_constraint({theta_idx: 1}, best_theta - 1e-7, 1)
    tier_profit = tiers["档位预测经营利润"].to_numpy(float)
    stage2_c = np.zeros(n_vars)
    stage2_c[:n_tiers] = -tier_profit
    result2 = milp(
        c=stage2_c,
        integrality=integrality,
        bounds=Bounds(lower, upper),
        constraints=linear_constraint(),
        options={"time_limit": 180, "mip_rel_gap": 1e-6},
    )
    if not result2.success:
        raise RuntimeError(f"第二阶段求解失败：{result2.message}")
    best_profit = float(np.dot(tier_profit, result2.x[:n_tiers]))

    # 第三阶段：在实际利润距最优值不超过 0.01 元时，
    # 最小化标准化价格偏离平方和。该指标不再视为经营成本。
    profit_floor = best_profit - PROFIT_TOLERANCE_YUAN
    add_constraint({j: tier_profit[j] for j in range(n_tiers)}, profit_floor, np.inf)
    stage3_c = np.zeros(n_vars)
    stage3_c[:n_tiers] = tiers["标准化价格偏离平方"].to_numpy(float)
    result3 = milp(
        c=stage3_c,
        integrality=integrality,
        bounds=Bounds(lower, upper),
        constraints=linear_constraint(),
        options={"time_limit": 180, "mip_rel_gap": 1e-6},
    )
    if not result3.success:
        raise RuntimeError(f"第三阶段求解失败：{result3.message}")

    x = result3.x
    strategy_rows: list[dict] = []
    for i, item in items.iterrows():
        js = np.flatnonzero(item_of_tier == i)
        selected_js = [int(j) for j in js if x[j] > 0.5]
        if not selected_js:
            continue
        j = selected_js[0]
        price = float(tiers.loc[j, "价格档位"])
        sales = float(tiers.loc[j, "价格决定的预测销量"])
        replenish = float(tiers.loc[j, "档位建议补货量"])
        cost = float(item["预测单品批发价(元/千克)"])
        price_deviation_index = float(tiers.loc[j, "标准化价格偏离平方"])
        revenue = price * sales
        purchase_cost = cost * replenish
        profit = revenue - purchase_cost
        strategy_rows.append(
            {
                "单品编码": item["单品编码"],
                "单品名称": item["单品名称"],
                "分类名称": item["分类名称"],
                "7月1日建议补货量(千克)": replenish,
                "7月1日建议售价(元/千克)": price,
                "预测销量(千克)": sales,
                "预测销售额(元)": revenue,
                "预测单品批发价(元/千克)": cost,
                "预测采购成本(元)": purchase_cost,
                "预测经营利润(元)": profit,
                "标准化价格偏离平方": price_deviation_index,
                "参考售价(元/千克)": item["参考售价(元/千克)"],
                "建议售价偏离率": price / float(item["参考售价(元/千克)"]) - 1,
                "建议成本加成率": price / cost - 1,
                "价格弹性β": item["价格弹性β"],
                "弹性P值": item["P值"],
                "弹性95%CI下限": item["95%CI下限"],
                "弹性95%CI上限": item["95%CI上限"],
                "单品损耗率": item["单品损耗率"],
                "近七日品类内份额": item["近七日品类内份额"],
                "有销量天数": item["有销量天数"],
            }
        )
    strategy = pd.DataFrame(strategy_rows)
    strategy["分类名称"] = pd.Categorical(strategy["分类名称"], CATEGORIES, ordered=True)
    strategy = strategy.sort_values(["分类名称", "7月1日建议补货量(千克)"], ascending=[True, False]).reset_index(drop=True)

    summary = (
        strategy.groupby("分类名称", observed=True)
        .agg(
            入选单品数=("单品编码", "count"),
            品类补货总量_kg=("7月1日建议补货量(千克)", "sum"),
            预测销量_kg=("预测销量(千克)", "sum"),
            预测销售额_元=("预测销售额(元)", "sum"),
            预测采购成本_元=("预测采购成本(元)", "sum"),
            预测经营利润_元=("预测经营利润(元)", "sum"),
        )
        .reset_index()
    )
    target_map = items.groupby("分类名称", observed=True)["品类目标需求(千克)"].first()
    summary["品类目标需求_kg"] = summary["分类名称"].map(target_map).astype(float)
    summary["需求满足率"] = summary["预测销量_kg"] / summary["品类目标需求_kg"]
    summary["需求缺口_kg"] = (summary["品类目标需求_kg"] - summary["预测销量_kg"]).clip(lower=0)

    diagnostics = {
        "第一阶段最大化最低品类满足率": best_theta,
        "第一阶段求解信息": result1.message,
        "第二阶段最优真实利润(元)": best_profit,
        "第二阶段求解信息": result2.message,
        "第三阶段利润容差(元)": PROFIT_TOLERANCE_YUAN,
        "第三阶段求解信息": result3.message,
        "第三阶段MIP相对间隙": getattr(result3, "mip_gap", np.nan),
        "第三阶段分支定界节点数": getattr(result3, "mip_node_count", np.nan),
        "最终标准化价格偏离平方和": strategy["标准化价格偏离平方"].sum(),
        "入选单品总数": len(strategy),
        "预测总销量(千克)": strategy["预测销量(千克)"].sum(),
        "建议总补货量(千克)": strategy["7月1日建议补货量(千克)"].sum(),
        "预测总销售额(元)": strategy["预测销售额(元)"].sum(),
        "预测总经营利润(元)": strategy["预测经营利润(元)"].sum(),
    }
    return strategy, summary, diagnostics


def run_sensitivity(
    items: pd.DataFrame,
    base_strategy: pd.DataFrame,
    base_summary: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """对无亏损约束做对照，并对需求、成本、损耗、弹性和价格带做扰动。"""
    rows: list[dict] = []

    def append_result(kind: str, name: str, strategy: pd.DataFrame, summary: pd.DataFrame) -> None:
        rows.append(
            {
                "类型": kind,
                "情景": name,
                "入选单品数": len(strategy),
                "总补货量(千克)": strategy["7月1日建议补货量(千克)"].sum(),
                "平均需求满足率": summary["需求满足率"].mean(),
                "最低品类需求满足率": summary["需求满足率"].min(),
                "预测总经营利润(元)": strategy["预测经营利润(元)"].sum(),
            }
        )

    append_result("基准", "基准情景", base_strategy, base_summary)

    # 约束对照：只保留单品预测经营利润非负的价格档位，
    # 等价于对每个入选单品施加 Pi_i >= 0。
    profitable_tiers = make_price_tiers(items)
    profitable_tiers = profitable_tiers[
        profitable_tiers["档位预测经营利润"] >= -1e-9
    ].reset_index(drop=True)
    profitable_tiers["档位序号"] = np.arange(len(profitable_tiers))
    no_loss_strategy, no_loss_summary, _ = solve_three_stage(items, profitable_tiers)
    append_result("约束对照", "入选单品利润均非负", no_loss_strategy, no_loss_summary)

    scenarios: list[tuple[str, pd.DataFrame]] = []

    for factor, label in [(0.95, "市场需求-5%"), (1.05, "市场需求+5%")]:
        scenario = items.copy()
        scenario["品类目标需求(千克)"] *= factor
        scenario["单品基准需求(千克)"] *= factor
        scenarios.append((label, scenario))

    for factor, label in [(0.95, "批发价格-5%"), (1.05, "批发价格+5%")]:
        scenario = items.copy()
        scenario["预测单品批发价(元/千克)"] *= factor
        scenarios.append((label, scenario))

    for factor, label in [(0.90, "价格弹性绝对值-10%"), (1.10, "价格弹性绝对值+10%")]:
        scenario = items.copy()
        scenario["价格弹性β"] *= factor
        scenarios.append((label, scenario))

    for factor, label in [(0.90, "损耗率-10%"), (1.10, "损耗率+10%")]:
        scenario = items.copy()
        scenario["单品损耗率"] = (scenario["单品损耗率"] * factor).clip(0, 0.50)
        scenarios.append((label, scenario))

    for label, scenario_items in scenarios:
        scenario_tiers = make_price_tiers(scenario_items)
        scenario_strategy, scenario_summary, _ = solve_three_stage(scenario_items, scenario_tiers)
        append_result("敏感性", label, scenario_strategy, scenario_summary)

    price_band_scenarios = [
        ("显著弹性品类价格带±10%", 0.10, INSIGNIFICANT_PRICE_BAND),
        ("显著弹性品类价格带±14%", 0.14, INSIGNIFICANT_PRICE_BAND),
        ("茄类价格带±3%", SIGNIFICANT_PRICE_BAND, 0.03),
        ("茄类价格带±7%", SIGNIFICANT_PRICE_BAND, 0.07),
    ]
    for label, significant_band, insignificant_band in price_band_scenarios:
        scenario_tiers = make_price_tiers(
            items,
            significant_band=significant_band,
            insignificant_band=insignificant_band,
        )
        scenario_strategy, scenario_summary, _ = solve_three_stage(items, scenario_tiers)
        append_result("价格区间敏感性", label, scenario_strategy, scenario_summary)

    result = pd.DataFrame(rows)
    base_profit = float(result.loc[result["情景"] == "基准情景", "预测总经营利润(元)"].iloc[0])
    result["相对基准利润变化率"] = result["预测总经营利润(元)"] / base_profit - 1
    return result, no_loss_strategy, no_loss_summary


def beautify_excel(path: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(path)
    fill = PatternFill("solid", fgColor="4472C4")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = fill
            cell.font = Font(name="Microsoft YaHei", color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in ws.columns:
            values = [str(c.value) if c.value is not None else "" for c in col[:200]]
            width = min(max(max(map(len, values)) + 2, 10), 28)
            ws.column_dimensions[col[0].column_letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Microsoft YaHei", size=10)
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    wb.save(path)


def save_figure(fig: plt.Figure, filename: str) -> None:
    fig.savefig(OUTPUT_DIR / filename, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def plot_strategy(strategy: pd.DataFrame) -> None:
    plot_df = strategy.sort_values(["分类名称", "7月1日建议补货量(千克)"], ascending=[False, True]).reset_index(drop=True)
    y = np.arange(len(plot_df))
    colors = plot_df["分类名称"].map(CATEGORY_COLORS).to_list()
    handles = [Line2D([0], [0], marker="o", color="w", markerfacecolor=CATEGORY_COLORS[c], markersize=8, label=c) for c in CATEGORIES]

    # 图 1-1：单品补货量
    fig1, ax1 = plt.subplots(figsize=(11, 11), facecolor="white")
    bars = ax1.barh(y, plot_df["7月1日建议补货量(千克)"], color=colors, alpha=0.82, edgecolor="white")
    ax1.set_yticks(y, plot_df["单品名称"])
    ax1.set_xlabel("建议补货量（千克）")
    ax1.set_title("2023年7月1日入选单品补货策略", fontsize=18, pad=16)
    ax1.grid(axis="x", color="#D8D8D8", alpha=0.55)
    ax1.set_axisbelow(True)
    ax1.bar_label(bars, fmt="%.2f", padding=3, fontsize=8, color="#555555")
    ax1.margins(x=0.10)
    ax1.legend(handles=handles, loc="lower center", ncol=6, frameon=False, bbox_to_anchor=(0.5, -0.08))
    fig1.subplots_adjust(left=0.25, right=0.97, top=0.94, bottom=0.10)
    save_figure(fig1, "图1-1_7月1日单品补货策略.png")

    # 图 1-2：单品定价
    fig2, ax2 = plt.subplots(figsize=(10, 11), facecolor="white")
    reference_prices = plot_df["参考售价(元/千克)"].to_numpy(float)
    suggested_prices = plot_df["7月1日建议售价(元/千克)"].to_numpy(float)
    for yi, ref_price, suggested_price in zip(y, reference_prices, suggested_prices):
        ax2.plot([ref_price, suggested_price], [yi, yi], color="#B8B8B8", linewidth=1.15, alpha=0.80, zorder=1)
    ax2.scatter(reference_prices, y, color="#777777", marker="|", s=95, linewidths=1.5, alpha=0.85, zorder=2)
    ax2.scatter(plot_df["7月1日建议售价(元/千克)"], y, c=colors, s=58, alpha=0.88, edgecolors="white", linewidths=0.6)
    for yi, price in zip(y, plot_df["7月1日建议售价(元/千克)"]):
        ax2.text(price, yi + 0.17, f"{price:.2f}", ha="center", va="bottom", fontsize=8, color="#555555")
    ax2.set_yticks(y, plot_df["单品名称"])
    ax2.set_xlabel("建议售价（元/千克）")
    ax2.set_title("2023年7月1日入选单品定价策略", fontsize=18, pad=16)
    ax2.grid(axis="x", color="#D8D8D8", alpha=0.55)
    ax2.set_axisbelow(True)
    price_handles = [Line2D([0], [0], marker="|", color="#777777", markersize=10, linestyle="None", label="近期参考售价")]
    ax2.legend(handles=price_handles + handles, loc="lower center", ncol=4, frameon=False, bbox_to_anchor=(0.5, -0.10))
    fig2.subplots_adjust(left=0.28, right=0.97, top=0.94, bottom=0.10)
    save_figure(fig2, "图1-2_7月1日单品定价策略.png")


def plot_category_summary(summary: pd.DataFrame) -> None:
    x = np.arange(len(summary))
    colors = [CATEGORY_COLORS[str(c)] for c in summary["分类名称"]]

    # 图 2-1：品类需求满足情况
    fig1, ax1 = plt.subplots(figsize=(12, 6.8), facecolor="white")
    width = 0.34
    ax1.bar(x - width / 2, summary["品类目标需求_kg"], width, color="#cee5df", alpha=0.8, label="目标需求")
    ax1.bar(x + width / 2, summary["预测销量_kg"], width, color=["#ead9ce","#afc1d1","#f5f0da","#eee8d9","#cbdab6","#ebd8d0"], alpha=0.86, label="优化后预测销量")
    ax1.set_ylabel("销量（千克）")
    ax1.set_xticks(x, summary["分类名称"])
    ax1.set_title("各品类市场需求满足情况", fontsize=18, pad=16)
    ax1.grid(axis="y", color="#D8D8D8", alpha=0.50)
    ax1.legend(frameon=False, ncol=2)
    for i, rate in enumerate(summary["需求满足率"]):
        ax1.text(i, max(summary.loc[i, "品类目标需求_kg"], summary.loc[i, "预测销量_kg"]) * 1.025, f"{rate:.1%}", ha="center", fontsize=10)
    fig1.tight_layout()
    save_figure(fig1, "图2-1_品类需求满足情况.png")

    # 图 2-2：各品类预测经营利润
    fig2, ax2 = plt.subplots(figsize=(11, 6.8), facecolor="white")
    ax2.bar(x, summary["预测经营利润_元"], color=["#e2b8aa","#e0d2cc","#c6d4d4","#ebd9d9","#b2bbd1","#cae3dd"], alpha=0.86, width=0.58)
    ax2.set_ylabel("预测经营利润（元）")
    ax2.set_xticks(x, summary["分类名称"])
    ax2.set_title("各品类预测经营利润", fontsize=18, pad=16)
    ax2.grid(axis="y", color="#D8D8D8", alpha=0.50)
    for i, value in enumerate(summary["预测经营利润_元"]):
        ax2.text(i, value, f"{value:.1f}", ha="center", va="bottom", fontsize=10)
    fig2.tight_layout()
    save_figure(fig2, "图2-2_各品类预测经营利润.png")



def plot_profit_risk(strategy: pd.DataFrame) -> None:
    fig, ax = plt.subplots(figsize=(12, 7.5), facecolor="white")
    for category in CATEGORIES:
        temp = strategy[strategy["分类名称"] == category]
        if temp.empty:
            continue
        sizes = 45 + 3.2 * temp["7月1日建议补货量(千克)"]
        ax.scatter(
            temp["单品损耗率"] * 100,
            temp["预测经营利润(元)"],
            s=sizes,
            color=CATEGORY_COLORS[category],
            alpha=0.76,
            edgecolors="white",
            linewidths=0.8,
            label=category,
        )
        for _, row in temp.nlargest(2, "预测经营利润(元)").iterrows():
            ax.annotate(row["单品名称"], (row["单品损耗率"] * 100, row["预测经营利润(元)"]), xytext=(5, 5), textcoords="offset points", fontsize=9)
    ax.axhline(0, color="#888888", linewidth=0.9)
    ax.set_xlabel("单品损耗率（%）")
    ax.set_ylabel("预测经营利润（元）")
    ax.set_title("入选单品的损耗风险与预测利润\n（气泡大小表示建议补货量）", fontsize=17)
    ax.grid(color="#D8D8D8", alpha=0.48)
    ax.legend(frameon=False, ncol=3, loc="best")
    fig.tight_layout()
    save_figure(fig, "图3_单品损耗风险与预测利润.png")
def plot_sensitivity_analysis(sensitivity: pd.DataFrame) -> None:
    """
    绘制第三问整体模型的敏感性分析图。
    正文推荐使用图4-1；图4-2可作为补充或放入附录。
    """
    sens = sensitivity[
        sensitivity["类型"].isin(["敏感性", "价格区间敏感性"])
    ].copy()

    if sens.empty:
        warnings.warn("敏感性分析结果为空，未生成敏感性图。")
        return

    # 固定论文展示顺序
    scenario_order = [
        "市场需求-5%",
        "市场需求+5%",
        "批发价格-5%",
        "批发价格+5%",
        "价格弹性绝对值-10%",
        "价格弹性绝对值+10%",
        "损耗率-10%",
        "损耗率+10%",
        "显著弹性品类价格带±10%",
        "显著弹性品类价格带±14%",
        "茄类价格带±3%",
        "茄类价格带±7%",
    ]

    sens["情景"] = pd.Categorical(
        sens["情景"],
        categories=scenario_order,
        ordered=True
    )

    sens = (
        sens.sort_values("情景")
        .dropna(subset=["情景"])
        .reset_index(drop=True)
    )

    # --------------------------------------------------------
    # 参数组及低饱和配色
    # --------------------------------------------------------
    def parameter_group(name: str) -> str:
        if name.startswith("市场需求"):
            return "市场需求"
        if name.startswith("批发价格"):
            return "批发价格"
        if name.startswith("价格弹性"):
            return "价格弹性"
        if name.startswith("损耗率"):
            return "损耗率"
        if name.startswith("显著弹性品类价格带"):
            return "显著弹性品类价格带"
        if name.startswith("茄类价格带"):
            return "茄类价格带"
        return "其他"

    group_colors = {
        "市场需求": "#8FAED3",
        "批发价格": "#D8AD79",
        "价格弹性": "#A99BC5",
        "损耗率": "#8DB79E",
        "显著弹性品类价格带": "#D89591",
        "茄类价格带": "#B29A80",
        "其他": "#AEB4BB",
    }

    sens["参数组"] = (
        sens["情景"]
        .astype(str)
        .map(parameter_group)
    )

    colors = (
        sens["参数组"]
        .map(group_colors)
        .tolist()
    )

    # ========================================================
    # 图4-1：关键参数扰动对预测经营利润的影响
    # ========================================================

    profit_change = pd.to_numeric(
        sens["相对基准利润变化率"],
        errors="coerce"
    ).to_numpy(float) * 100

    y = np.arange(len(sens))

    fig1, ax1 = plt.subplots(
        figsize=(11.5, 7.4),
        facecolor="white"
    )

    # 基准线
    ax1.axvline(
        0,
        color="#8A8A8A",
        linewidth=1.0,
        linestyle="--",
        alpha=0.75,
        zorder=1,
    )

    # 连接基准值与扰动结果
    for yi, value, color in zip(
        y,
        profit_change,
        colors
    ):
        ax1.plot(
            [0, value],
            [yi, yi],
            color=color,
            linewidth=3.0,
            alpha=0.55,
            solid_capstyle="round",
            zorder=2,
        )

        ax1.scatter(
            value,
            yi,
            s=95,
            color=color,
            alpha=0.88,
            edgecolors="white",
            linewidths=0.8,
            zorder=3,
        )

    # 数值标注
    span = max(
        float(
            np.nanmax(profit_change)
            -
            np.nanmin(profit_change)
        ),
        1.0
    )

    for yi, value in zip(
        y,
        profit_change
    ):
        offset = span * 0.018

        if value >= 0:
            x_text = value + offset
            ha = "left"
        else:
            x_text = value - offset
            ha = "right"

        ax1.text(
            x_text,
            yi,
            f"{value:+.1f}%",
            va="center",
            ha=ha,
            fontsize=10,
            color="#444444"
        )

    ax1.set_yticks(
        y,
        sens["情景"].astype(str)
    )

    ax1.invert_yaxis()

    ax1.set_xlabel(
        "相对基准利润变化率（%）"
    )

    ax1.set_ylabel(
        "扰动情景"
    )

    ax1.set_title(
        "关键参数扰动对预测经营利润的影响",
        fontsize=18,
        pad=16
    )

    ax1.grid(
        axis="x",
        color="#D8D8D8",
        linestyle="--",
        alpha=0.35
    )

    ax1.set_axisbelow(True)

    ax1.spines["top"].set_visible(False)
    ax1.spines["right"].set_visible(False)

    # 图例
    present_groups = list(
        dict.fromkeys(
            sens["参数组"].tolist()
        )
    )

    handles = [
        Line2D(
            [0],
            [0],
            marker="o",
            linestyle="None",
            markerfacecolor=group_colors[g],
            markeredgecolor="white",
            markersize=8,
            label=g,
        )
        for g in present_groups
    ]

    ax1.legend(
        handles=handles,
        frameon=False,
        ncol=3,
        loc="lower center",
        bbox_to_anchor=(0.5, -0.18)
    )

    fig1.subplots_adjust(
        left=0.30,
        right=0.96,
        top=0.91,
        bottom=0.19
    )

    save_figure(
        fig1,
        "图4-1_敏感性分析_利润变化.png"
    )

    # ========================================================
    # 图4-2：关键参数扰动对最低需求满足率的影响
    # ========================================================

    min_rate = pd.to_numeric(
        sens["最低品类需求满足率"],
        errors="coerce"
    ).to_numpy(float) * 100

    base_row = sensitivity.loc[
        sensitivity["情景"] == "基准情景"
    ]

    if not base_row.empty:
        base_rate = (
            float(
                base_row[
                    "最低品类需求满足率"
                ].iloc[0]
            )
            * 100
        )
    else:
        base_rate = float(
            np.nanmedian(min_rate)
        )

    fig2, ax2 = plt.subplots(
        figsize=(11.5, 7.4),
        facecolor="white"
    )

    ax2.axvline(
        base_rate,
        color="#8A8A8A",
        linewidth=1.0,
        linestyle="--",
        alpha=0.75,
        label=f"基准最低满足率 {base_rate:.2f}%",
        zorder=1,
    )

    for yi, value, color in zip(
        y,
        min_rate,
        colors
    ):
        ax2.plot(
            [base_rate, value],
            [yi, yi],
            color=color,
            linewidth=3.0,
            alpha=0.55,
            solid_capstyle="round",
            zorder=2,
        )

        ax2.scatter(
            value,
            yi,
            s=95,
            color=color,
            alpha=0.88,
            edgecolors="white",
            linewidths=0.8,
            zorder=3,
        )

        ax2.text(
            value + 0.03,
            yi,
            f"{value:.2f}%",
            va="center",
            ha="left",
            fontsize=10,
            color="#444444",
        )

    ax2.set_yticks(
        y,
        sens["情景"].astype(str)
    )

    ax2.invert_yaxis()

    ax2.set_xlabel(
        "最低品类需求满足率（%）"
    )

    ax2.set_ylabel(
        "扰动情景"
    )

    ax2.set_title(
        "关键参数扰动对最低需求满足率的影响",
        fontsize=18,
        pad=16
    )

    ax2.grid(
        axis="x",
        color="#D8D8D8",
        linestyle="--",
        alpha=0.35
    )

    ax2.set_axisbelow(True)

    ax2.spines["top"].set_visible(False)
    ax2.spines["right"].set_visible(False)

    ax2.legend(
        frameon=False,
        loc="lower right"
    )

    fig2.subplots_adjust(
        left=0.30,
        right=0.96,
        top=0.91,
        bottom=0.10
    )

    save_figure(
        fig2,
        "图4-2_敏感性分析_最低满足率.png"
    )

def write_model_note(strategy: pd.DataFrame, summary: pd.DataFrame, diagnostics: dict, sensitivity: pd.DataFrame) -> None:
    total_replenish = strategy["7月1日建议补货量(千克)"].sum()
    total_profit = strategy["预测经营利润(元)"].sum()
    min_rate = summary["需求满足率"].min()
    avg_rate = summary["需求满足率"].mean()
    negative_categories = summary.loc[summary["预测经营利润_元"] < 0, ["分类名称", "预测经营利润_元"]]
    negative_item_count = int((strategy["预测经营利润(元)"] < 0).sum())
    no_loss_row = sensitivity.loc[sensitivity["情景"] == "入选单品利润均非负"].iloc[0]
    significant_10_row = sensitivity.loc[
        sensitivity["情景"] == "显著弹性品类价格带±10%"
    ].iloc[0]
    significant_14_row = sensitivity.loc[
        sensitivity["情景"] == "显著弹性品类价格带±14%"
    ].iloc[0]
    eggplant_3_row = sensitivity.loc[sensitivity["情景"] == "茄类价格带±3%"].iloc[0]
    eggplant_7_row = sensitivity.loc[sensitivity["情景"] == "茄类价格带±7%"].iloc[0]
    summary_rows = [
        f"| {row['分类名称']} | {int(row['入选单品数'])} | {row['品类目标需求_kg']:.2f} | "
        f"{row['品类补货总量_kg']:.2f} | {row['需求满足率']:.2%} | {row['预测经营利润_元']:.2f} |"
        for _, row in summary.iterrows()
    ]
    sensitivity_rows = [
        f"| {row['类型']} | {row['情景']} | {int(row['入选单品数'])} | {row['总补货量(千克)']:.2f} | "
        f"{row['平均需求满足率']:.2%} | {row['最低品类需求满足率']:.2%} | "
        f"{row['预测总经营利润(元)']:.2f} | {row['相对基准利润变化率']:.2%} |"
        for _, row in sensitivity.iterrows()
    ]
    lines = [
        "# 第三问：基于序贯混合整数规划的单品补货与定价",
        "",
        "## 1. 问题分析",
        "",
        "本问需在商超空间、最小陈列量和日内损耗的共同约束下，同时决定单品选择、补货量和售价。价格改变单品可销量，品种选择影响各品类需求的满足度，补货量过高又会造成损耗成本。因此，将该问题建模为序贯混合整数线性规划。",
        "",
        "## 2. 模型假设",
        "",
        "1. 6 月 24—30 日有销量的单品视为 7 月 1 日可采购候选集，共 49 个。",
        "2. 品类总需求使用第二问在参考价格水平下预测的 7 月 1 日基准需求，不使用问题二最优价格对应的优化后销量；近 7 日品类内销量份额在短期内基本稳定。",
        "3. 单品基准需求由品类需求按近 7 日销量份额分解；历史最大日销量仅用于数据诊断，不代替需求函数的基准量。",
        "4. 由于单品层面的日数据较少，同一品类的不同单品共用该品类的价格弹性。茄类弹性的 P 值为 0.458，未达 5% 显著水平，因此将茄类售价波动带收窄为参考售价附近 ±5%；其他品类为 ±12%。",
        "5. 单品近期损耗率在 7 月 1 日保持不变，并设置 5% 安全库存。",
        "6. 未显式建模单品之间的需求转移；某单品未入选时，其需求不会自动转移到同品类其他单品。",
        "",
        "## 3. 参数估计",
        "",
        "令 $Q_k$ 为问题二在品类参考价格水平下得到的 7 月 1 日基准需求预测，单品 $i$ 近 7 日销量为 $S_i$，则单品基准需求为 $d_i^0=Q_kS_i/\\sum_{h\\in k}S_h$。因此 $Q_k$ 与后续价格弹性公式均以参考价格为共同基准，不会重复计入问题二最优价格的需求影响。单品参考售价 $p_i^0$ 取近 7 日销额与销量之比。",
        "",
        f"设 $c_{{it}}$ 为决策日前第 $t$ 日的单品批发价，$a_{{it}}=\\exp(-{COST_DECAY_RATE:.2f}\\Delta t)$ 为时间衰减权重，则近 14 日加权成本为 $\\bar c_i=\\sum_ta_{{it}}c_{{it}}/\\sum_ta_{{it}}$。令 $w_i=S_i/\\sum_{{h\\in k}}S_h$，$\\widehat C_k$ 为问题二预测的 7 月 1 日品类批发价，则品类校正系数为 $\\alpha_k=\\widehat C_k/\\sum_{{i\\in k}}w_i\\bar c_i$，最终单品成本为 $c_i=\\alpha_k\\bar c_i$。",
        "",
        "## 4. 决策变量",
        "",
        "- $z_{ij}\u2208\\{0,1\\}$：单品 $i$ 是否选择价格档位 $j$；$x_i=\\sum_jz_{ij}$ 表示单品是否入选。",
        "- $\\theta\u2208[0,1]$：六大品类中的最低需求满足率。",
        "- $D_{ij}$ 和 $q_{ij}$ 分别是价格档位 $j$ 唯一确定的预测销量和补货量，它们是预计算参数，不再是优化器可任意选择的变量。",
        "",
        "## 5. 序贯混合整数规划",
        "",
        "将每个单品的可行售价区间离散为 7 个等距档位，并补入参考售价档位。一旦选中档位 $j$，销量严格取 $D_{ij}=d_i^0(p_{ij}/p_i^0)^{\\beta_k}$，补货量严格取 $q_{ij}=\\max\\{2.5,1.05D_{ij}/(1-l_i)\\}$。因此商超只决定价格与品种，不直接决定顾客购买量。",
        "",
        "令 $P_{i,10%}$ 和 $P_{i,90%}$ 为近期售价分位数，$\\delta_i$ 对茄类取 5%、其他品类取 12%。最终价格下界与上界为 $p_i^L=\\max\\{(1-\\delta_i)p_i^0,\\min(P_{i,10%},p_i^0)\\}$ 和 $p_i^U=\\min\\{(1+\\delta_i)p_i^0,\\max(P_{i,90%},p_i^0)\\}$。",
        "",
        "第一阶段最大化六个品类中最差者的需求满足率：$\\max\\theta$，且对任意品类 $k$ 有 $\\theta Q_k\\le\\sum_{i\\in k,j}D_{ij}z_{ij}\\le Q_k$。第二阶段固定 $\\theta=\\theta^*$，直接最大化题目要求的真实经营利润：",
        "",
        "$$\\max \\Pi=\\sum_{i,j}(p_{ij}D_{ij}-c_iq_{ij})z_{ij}.$$",
        "",
        f"第三阶段在 $\\Pi\\ge\\Pi^*-{PROFIT_TOLERANCE_YUAN:.2f}$ 元的条件下，最小化标准化价格偏离 $\\sum_{{i,j}}[(p_{{ij}}-p_i^0)/\\sigma_i]^2z_{{ij}}$。其中 $\\sigma_i=\\max\\{{\\operatorname{{sd}}(P_{{it}}),0.05p_i^0\\}}$，即采用单品近期有效售价标准差，并设置参考售价 5% 的尺度下限，避免标准差过小导致惩罚失真。价格偏离仅是第三阶段的次级目标，不再被视为真实成本，因此第二阶段所得 $\\Pi^*$ 与题意的“收益最大”完全对应。",
        "",
        "品种数与价格唯一性约束为 $27\\le\\sum_ix_i\\le33$和 $\\sum_jz_{ij}\\le1$。不再额外规定“每类至少 2 个单品”；各品类是否入选及其数量由最低满足率目标内生决定。",
        "",
        "## 6. 求解结果",
        "",
        f"模型最终选择 **{len(strategy)}** 个单品，建议总补货量为 **{total_replenish:.2f} kg**，预测当日经营利润为 **{total_profit:.2f} 元**。六大品类平均需求满足率为 **{avg_rate:.2%}**，最低品类满足率为 **{min_rate:.2%}**。",
        "",
        "| 品类 | 入选数 | 目标需求/kg | 补货量/kg | 需求满足率 | 预测利润/元 |",
        "|---|---:|---:|---:|---:|---:|",
        *summary_rows,
        "",
        f"由于模型未设置额外 SKU 固定经营成本，增加品种可改善最差品类满足率，因此最终使用题目允许的上限 {len(strategy)} 种。这是最低满足率优先下的内生结果，而非人为规定。具体单品补货量和售价见 Excel 的“最优单品策略”工作表。",
        "",
        (
            "在“品类最低满足率优先”的序贯目标下，"
            + "、".join(f"{row['分类名称']}（{row['预测经营利润_元']:.2f}元）" for _, row in negative_categories.iterrows())
            + "出现品类层面的预测亏损，表明服务公平性与品类利润之间存在交易。但优化对象是商超整体收益，故允许品类间交叉补贴，且总利润仍为正。"
            if not negative_categories.empty
            else "各品类预测经营利润均为正，无需依赖品类间交叉补贴。"
        ),
        "",
        f"主方案中共有 {negative_item_count} 个入选单品的预测利润为负。将约束 $\\Pi_i\\ge0$ 加入后，最低品类满足率为 {no_loss_row['最低品类需求满足率']:.2%}，总利润为 {no_loss_row['预测总经营利润(元)']:.2f} 元。该对照用于量化“无亏损经营”与“品类服务水平”之间的交易。",
        "",
        "## 7. 敏感性分析",
        "",
        "| 类型 | 情景 | 入选数 | 补货量/kg | 平均满足率 | 最低满足率 | 预测利润/元 | 利润变化率 |",
        "|---|---|---:|---:|---:|---:|---:|---:|",
        *sensitivity_rows,
        "",
        "敏感性分析分别对市场需求和批发价格作 ±5% 扰动，对价格弹性绝对值和损耗率作 ±10% 扰动；同时将显著弹性品类的允许价格带由基准 ±12% 调整为 ±10% 和 ±14%，将茄类价格带由基准 ±5% 调整为 ±3% 和 ±7%。每个情景均重新执行完整的三阶段选品与定价优化。",
        "",
        f"显著弹性品类的价格带由 ±10% 放宽到 ±14% 时，最低需求满足率由 {significant_10_row['最低品类需求满足率']:.2%} 提高到 {significant_14_row['最低品类需求满足率']:.2%}，但利润由 {significant_10_row['预测总经营利润(元)']:.2f} 元降至 {significant_14_row['预测总经营利润(元)']:.2f} 元。这是序贯目标下服务水平优先于利润的结果：更宽价格带提高了第一阶段可达到的公平服务水平，第二阶段随后在更高服务约束下求利润最优。茄类价格带从 ±3% 到 ±7% 时，最低满足率均为 {eggplant_3_row['最低品类需求满足率']:.2%}，利润仅由 {eggplant_3_row['预测总经营利润(元)']:.2f} 元变为 {eggplant_7_row['预测总经营利润(元)']:.2f} 元，说明结论对茄类价格带设定相对稳定。",
        "",
        "## 8. 模型局限",
        "",
        "单品基准需求依赖近 7 日品类内销量份额，同品类单品共用价格弹性，且未显式刻画单品间替代效应。如能获得更长时段的单品价格试验和缺货替代数据，可进一步建立单品层面的交叉价格弹性模型。",
        "",
        f"本模型将价格离散为 {PRICE_GRID_SIZE} 个等距档位并补入参考价档位，因此求解器返回的是“给定价格离散精度下的混合整数规划全局最优解”，不宣称为原连续定价问题的全局最优解。",
        "",
        f"第三阶段求解器状态：{diagnostics['第三阶段求解信息']}",
    ]
    (OUTPUT_DIR / "第三问_建模说明.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    require_inputs()
    font_name = configure_chinese_font()
    print(f"论文图字体：{font_name}")
    candidates, daily, wholesale, category_plan, item_loss, elasticity = read_inputs()
    items = build_item_parameters(candidates, daily, wholesale, category_plan, item_loss, elasticity)
    tiers = make_price_tiers(items)
    strategy, summary, diagnostics = solve_three_stage(items, tiers)
    sensitivity, no_loss_strategy, no_loss_summary = run_sensitivity(items, strategy, summary)

    diagnostic_df = pd.DataFrame({"指标": list(diagnostics.keys()), "数值": list(diagnostics.values())})
    excel_path = OUTPUT_DIR / "第三问_单品补货与定价结果.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        strategy.to_excel(writer, sheet_name="最优单品策略", index=False)
        summary.to_excel(writer, sheet_name="品类汇总", index=False)
        items.drop(columns=["原始成本份额乘积"], errors="ignore").to_excel(writer, sheet_name="候选单品参数", index=False)
        sensitivity.to_excel(writer, sheet_name="敏感性与约束对照", index=False)
        no_loss_strategy.to_excel(writer, sheet_name="无亏损单品对照策略", index=False)
        no_loss_summary.to_excel(writer, sheet_name="无亏损对照品类汇总", index=False)
        diagnostic_df.to_excel(writer, sheet_name="求解诊断", index=False)
    beautify_excel(excel_path)

    plot_strategy(strategy)
    plot_category_summary(summary)
    plot_profit_risk(strategy)
    write_model_note(strategy, summary, diagnostics, sensitivity)
    # 新增：敏感性分析绘图
    plot_sensitivity_analysis(sensitivity)

    write_model_note(
        strategy,
        summary,
        diagnostics,
        sensitivity
    )
    print("\n" + "=" * 78)
    print("第三问求解完成")
    print("=" * 78)
    print(f"候选单品数：{len(items)}")
    print(f"入选单品数：{len(strategy)}")
    print(f"建议总补货量：{strategy['7月1日建议补货量(千克)'].sum():.3f} kg")
    print(f"预测总经营利润：{strategy['预测经营利润(元)'].sum():.2f} 元")
    print(f"平均需求满足率：{summary['需求满足率'].mean():.2%}")
    print(f"结果目录：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
