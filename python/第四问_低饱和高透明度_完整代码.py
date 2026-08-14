from __future__ import annotations

from pathlib import Path
from typing import Any

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import font_manager, patches
from matplotlib.colors import LinearSegmentedColormap


# =====================================================================
# 1. 路径与模型参数
# =====================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
OUTPUT_DIR = SCRIPT_DIR / "第四问新"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

FONT_FILE = PROJECT_DIR / "fonts" / "msyh.ttc"
QUALITY_FILE = SCRIPT_DIR / "数据质量检查表.xlsx"

INPUT_SPECS = [
    (PROJECT_DIR / "附件1.xlsx", "Sheet1", "附件1：单品信息"),
    (PROJECT_DIR / "附件2.xlsx", "Sheet1", "附件2：销售流水"),
    (PROJECT_DIR / "附件3.xlsx", "Sheet1", "附件3：批发价格"),
    (PROJECT_DIR / "附件4.xlsx", "Sheet1", "附件4：单品损耗率"),
]

QUESTION_LABELS = ["问题1", "问题2", "问题3"]
BASE_WEIGHTS = {"影响": 0.40, "缺口": 0.30, "可采集性": 0.20, "紧迫性": 0.10}
MONTE_CARLO_RUNS = 10_000
MONTE_CARLO_SEED = 20260814
WEIGHT_PERTURBATION = 0.20
SCORE_PERTURBATION = 0.50

# =====================================================================
# 论文图统一低饱和、高透明度配色
# =====================================================================
# 色彩原则：
# 1. 低饱和：避免大红、大蓝等高刺激颜色；
# 2. 高透明：主体填充保持 0.50~0.68 的透明度；
# 3. 自然过渡：以雾蓝、灰绿、燕麦黄、豆沙粉、浅薰衣草为主；
# 4. 统一风格：四张图及 Excel 批次颜色保持一致。

PAPER_BG = "#FBFAF7"       # 温和米白背景
TEXT_DARK = "#4E4A46"      # 深灰褐文字
TEXT_MID = "#6F6963"       # 中灰褐文字
GRID_COLOR = "#D8D4CD"     # 淡灰米色网格
EDGE_COLOR = "#8E8881"     # 柔和边框

BATCH_COLORS = {
    "第一批—立即建设": "#C79A9A",   # 低饱和豆沙粉
    "第二批—近期建设": "#D8BE8D",   # 燕麦金
    "第三批—持续完善": "#9FB8C9",   # 雾霾蓝
}

FLOW_COLORS = [
    "#A8BED0",  # 雾蓝
    "#D8B995",  # 杏仁沙
    "#A7C0B2",  # 鼠尾草绿
    "#B7AEC8",  # 浅薰衣草
]

MC_BAR_COLOR = "#8FAFC3"   # 蒙特卡洛柱图：灰蓝
MC_POINT_COLOR = "#C59A9A" # 蒙特卡洛误差线：灰粉

# 影响矩阵：米白 -> 灰绿 -> 雾蓝，变化柔和
IMPACT_CMAP = LinearSegmentedColormap.from_list(
    "soft_natural",
    ["#F7F3EA", "#E5E4D3", "#D2DDD2", "#B8CFC3", "#9BBDB6", "#7FA5AE"],
    N=256,
)


def configure_chinese_font() -> str:
    """显式注册项目中的微软雅黑，保证图片可直接用于论文。"""
    if not FONT_FILE.exists():
        raise FileNotFoundError(f"未找到微软雅黑字体：{FONT_FILE}")
    font_manager.fontManager.addfont(str(FONT_FILE))
    font_name = font_manager.FontProperties(fname=str(FONT_FILE)).get_name()
    plt.rcParams.update(
        {
            "font.family": "sans-serif",
            "font.sans-serif": [font_name],
            "axes.unicode_minus": False,
            "figure.dpi": 120,
            "savefig.dpi": 320,
            "axes.edgecolor": EDGE_COLOR,
            "axes.labelcolor": TEXT_DARK,
            "xtick.color": TEXT_MID,
            "ytick.color": TEXT_MID,
            "text.color": TEXT_DARK,
            "figure.facecolor": PAPER_BG,
            "axes.facecolor": PAPER_BG,
        }
    )
    return font_name


def require_inputs() -> None:
    required = [QUALITY_FILE, FONT_FILE] + [path for path, _, _ in INPUT_SPECS]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("缺少输入文件：" + "、".join(missing))


# =====================================================================
# 2. 数据需求字典
# =====================================================================

# 每个概念对应一组可能的字段别名。覆盖率只统计“已有且可直接使用”的字段，
# 不把相近但口径不同的字段误认为目标数据已经具备。例如日批发价不能替代逐笔实际结算价，
# 历史销售量也不能替代采用策略后的执行反馈销量。
DATA_REQUIREMENTS: list[dict[str, Any]] = [
    {
        "数据主题": "库存、缺货与陈列状态",
        "字段概念": [
            ("期初库存", ["期初库存", "期初存量"]),
            ("入库量", ["入库量", "进货量", "到货量"]),
            ("期末库存", ["期末库存", "库存结余"]),
            ("缺货标志", ["缺货标志", "缺货状态", "断货标记"]),
            ("陈列数量", ["陈列数量", "陈列量", "货架库存"]),
            ("补货时间", ["补货时间", "上架时间"]),
        ],
        "采集频率": "实时/日度",
        "数据来源": "库存系统、电子秤、货架巡检",
        "问题1影响": 2,
        "问题2影响": 5,
        "问题3影响": 5,
        "可采集性评分": 4,
        "紧迫性评分": 5,
        "问题1帮助": "识别销量下降究竟来自真实需求变化还是缺货导致的观测偏差。",
        "问题2帮助": "提高品类需求预测，避免把缺货期间的低销量当成市场低需求。",
        "问题3帮助": "为单品补货量、陈列约束和需求满足率提供真实状态变量。",
        "核心作用": "纠正缺货截断需求，是需求预测和补货优化的基础数据。",
    },
    {
        "数据主题": "订货、到货、供应商与交期",
        "字段概念": [
            ("订单编号", ["订单编号", "采购订单号"]),
            ("供应商编码", ["供应商编码", "供应商ID"]),
            ("下单时间", ["下单时间", "采购时间"]),
            ("到货时间", ["到货时间", "收货时间"]),
            ("订货量", ["订货量", "采购量"]),
            ("实际结算价", ["实际结算价", "结算单价"]),
            ("最低起订量", ["最低起订量", "最小采购量"]),
        ],
        "采集频率": "每笔订单/到货时",
        "数据来源": "采购系统、供应商协同平台、收货验收单",
        "问题1影响": 1,
        "问题2影响": 4,
        "问题3影响": 5,
        "可采集性评分": 4,
        "紧迫性评分": 5,
        "问题1帮助": "补充供给侧解释变量，区分销量关联中的供应波动影响。",
        "问题2帮助": "改进批发价预测，并把交期和采购量约束纳入一周补货决策。",
        "问题3帮助": "将供应商、交期、最低起订量和真实采购成本纳入单品MILP。",
        "核心作用": "把估计成本和理想补货约束替换为真实采购约束。",
    },
    {
        "数据主题": "促销、折扣原因与价格变动",
        "字段概念": [
            ("挂牌价", ["挂牌价", "原价"]),
            ("成交售价", ["销售单价(元/千克)", "成交价"]),
            ("折扣标记", ["是否打折销售"]),
            ("促销活动编号", ["促销活动编号", "活动ID"]),
            ("折扣原因", ["折扣原因", "促销类型", "临期处理原因"]),
            ("价格生效时段", ["价格生效时段", "价格变更时间"]),
        ],
        "采集频率": "每次价格变更/交易时",
        "数据来源": "POS、促销管理系统、价签系统",
        "问题1影响": 4,
        "问题2影响": 5,
        "问题3影响": 5,
        "可采集性评分": 5,
        "紧迫性评分": 5,
        "问题1帮助": "区分单品销量相关性是共同促销造成，还是商品自身需求造成。",
        "问题2帮助": "降低促销混杂对价格弹性和成本加成率估计的影响。",
        "问题3帮助": "支持分时定价、折价销售和价格档位的更准确需求函数。",
        "核心作用": "建立可解释的价格—需求关系，避免把促销效果误当成价格弹性。",
    },
    {
        "数据主题": "逐笔报损、折价处理与损耗原因",
        "字段概念": [
            ("报损量", ["报损量", "废弃量"]),
            ("报损时间", ["报损时间", "处置时间"]),
            ("报损原因", ["报损原因", "损耗原因"]),
            ("折价处置量", ["折价处置量", "临期折价量"]),
            ("最终处置方式", ["最终处置方式", "处置方式"]),
        ],
        "采集频率": "每次处置时",
        "数据来源": "报损系统、临期处理记录、盘点记录",
        "问题1影响": 2,
        "问题2影响": 4,
        "问题3影响": 5,
        "可采集性评分": 3,
        "紧迫性评分": 5,
        "问题1帮助": "识别销量分布与损耗分布的共同波动关系。",
        "问题2帮助": "把损耗成本和折价收入纳入品类利润预测。",
        "问题3帮助": "将单品损耗率由静态均值改为随库存、时间和品相变化的损耗函数。",
        "核心作用": "直接提升损耗成本、折价收入和利润预测的真实性。",
    },
    {
        "数据主题": "质量、等级、规格、产地与新鲜度",
        "字段概念": [
            ("质量等级", ["质量等级", "商品等级"]),
            ("规格重量", ["规格重量", "包装规格"]),
            ("产地", ["产地", "供应产地"]),
            ("到货/采收日期", ["到货日期", "采收日期"]),
            ("新鲜度评分", ["新鲜度评分", "品相评分"]),
            ("保质期", ["保质期", "预计可售天数"]),
        ],
        "采集频率": "到货时/每日盘点",
        "数据来源": "收货验收、质检、图像识别、供应商批次",
        "问题1影响": 2,
        "问题2影响": 3,
        "问题3影响": 5,
        "可采集性评分": 3,
        "紧迫性评分": 4,
        "问题1帮助": "解释不同供应来源或规格造成的单品销量差异。",
        "问题2帮助": "提升不同等级商品的需求和价格分层预测。",
        "问题3帮助": "将质量差异、可售期和损耗风险纳入单品选品与定价。",
        "核心作用": "解释单品异质性，避免把不同品质商品强行视为同质单品。",
    },
    {
        "数据主题": "匿名顾客、订单与购物篮",
        "字段概念": [
            ("匿名顾客ID", ["匿名顾客ID", "会员ID"]),
            ("订单/购物篮编号", ["订单编号", "购物篮编号"]),
            ("购物篮商品序列", ["商品序列", "购买顺序"]),
            ("购买渠道", ["购买渠道", "门店/线上渠道"]),
        ],
        "采集频率": "每笔订单",
        "数据来源": "会员系统、POS订单、线上订单",
        "问题1影响": 5,
        "问题2影响": 4,
        "问题3影响": 5,
        "可采集性评分": 2,
        "紧迫性评分": 3,
        "问题1帮助": "从共现关系进一步识别替代、互补和购物篮关联。",
        "问题2帮助": "改善品类之间的联合需求预测和促销组合设计。",
        "问题3帮助": "在单品未入选时估计需求转移，改进单品需求分解。",
        "核心作用": "把相关关系提升为更接近消费行为的替代和互补关系。",
    },
    {
        "数据主题": "客流、天气、节假日与竞品价格",
        "字段概念": [
            ("门店客流", ["门店客流", "进店人数"]),
            ("天气温度", ["天气温度", "平均气温"]),
            ("降雨/极端天气", ["降雨量", "天气类型"]),
            ("节假日/节庆", ["节假日", "节庆活动"]),
            ("竞品价格", ["竞品价格", "周边门店价格"]),
        ],
        "采集频率": "小时/日度",
        "数据来源": "客流系统、气象接口、节假日日历、竞品采价",
        "问题1影响": 3,
        "问题2影响": 5,
        "问题3影响": 4,
        "可采集性评分": 4,
        "紧迫性评分": 4,
        "问题1帮助": "解释共同时间波动，避免把外部冲击误判为品类间关联。",
        "问题2帮助": "增强未来一周需求预测和动态定价的外生解释变量。",
        "问题3帮助": "帮助识别特定日期和客流情景下的单品需求变化。",
        "核心作用": "提高短期需求预测的情景适应性和可解释性。",
    },
    {
        "数据主题": "货架、冷链和补货作业约束",
        "字段概念": [
            ("货架容量", ["货架容量", "陈列容量"]),
            ("冷藏容量", ["冷藏容量", "冷库容量"]),
            ("货架位置", ["货架位置", "陈列区域"]),
            ("补货人力/耗时", ["补货人力", "补货耗时"]),
            ("设备可用时段", ["设备可用时段", "作业时间窗"]),
        ],
        "采集频率": "日度/变更时",
        "数据来源": "门店运营系统、货架巡检、作业记录",
        "问题1影响": 1,
        "问题2影响": 2,
        "问题3影响": 5,
        "可采集性评分": 4,
        "紧迫性评分": 4,
        "问题1帮助": "补充销售空间和陈列位置对销量差异的解释。",
        "问题2帮助": "将品类补货计划与门店可执行能力连接起来。",
        "问题3帮助": "把当前的SKU数量和最小陈列约束扩展为真实容量与作业约束。",
        "核心作用": "保证优化方案不仅数学可行，也能在门店现场执行。",
    },
    {
        "数据主题": "策略执行反馈与价格试验",
        "字段概念": [
            ("建议售价", ["建议售价", "策略售价"]),
            ("实际执行售价", ["实际执行售价", "执行售价"]),
            ("建议补货量", ["建议补货量", "计划补货量"]),
            ("实际到货/补货量", ["实际到货量", "实际补货量"]),
            ("执行后销量", ["执行后销量"]),
            ("执行后报损", ["执行后报损", "报损量"]),
            ("价格试验组标记", ["价格试验组", "A/B组标记"]),
        ],
        "采集频率": "每次决策/日度回收",
        "数据来源": "策略系统、POS、采购和报损系统",
        "问题1影响": 3,
        "问题2影响": 5,
        "问题3影响": 5,
        "可采集性评分": 3,
        "紧迫性评分": 5,
        "问题1帮助": "用执行结果验证历史关联关系是否具有稳定性。",
        "问题2帮助": "评估需求预测、定价和利润预测的误差并持续校准。",
        "问题3帮助": "检验单品补货定价策略，形成滚动更新和安全边界。",
        "核心作用": "形成模型闭环，使预测参数不再一次性固定。",
    },
]


def read_current_inventory() -> tuple[pd.DataFrame, set[str], dict[str, int]]:
    """读取所有附件的字段，并用现有质量表或实际小表统计记录数。"""
    quality = pd.read_excel(QUALITY_FILE)
    row_count_map: dict[str, int] = {}
    for _, row in quality.iterrows():
        if row["检查项目"] == "原始记录数":
            try:
                row_count_map[str(row["数据表"])] = int(row["检查结果"])
            except (TypeError, ValueError):
                continue

    rows: list[dict[str, Any]] = []
    all_fields: set[str] = set()
    for path, sheet, label in INPUT_SPECS:
        header = pd.read_excel(path, sheet_name=sheet, nrows=0)
        fields = [str(c) for c in header.columns]
        all_fields.update(fields)
        source_name = path.stem
        if source_name == "附件2":
            row_count = row_count_map.get("附件2", 0)
        else:
            row_count = int(len(pd.read_excel(path, sheet_name=sheet, usecols=[0])))
        rows.append(
            {
                "数据源": label,
                "文件": path.name,
                "工作表": sheet,
                "记录数": row_count,
                "字段数": len(fields),
                "已有字段": "、".join(fields),
            }
        )

    quality_rows = quality[["数据表", "检查项目", "检查结果", "备注"]].copy()
    quality_rows["数据源"] = "数据质量检查表"
    quality_rows["文件"] = QUALITY_FILE.name
    quality_rows["工作表"] = "Sheet1"
    quality_rows["记录数"] = np.nan
    quality_rows["字段数"] = np.nan
    quality_rows["已有字段"] = "质量检查指标：" + quality_rows["检查项目"].astype(str)
    quality_rows = quality_rows[rows[0].keys()]
    inventory_df = pd.concat([pd.DataFrame(rows), quality_rows], ignore_index=True)
    return inventory_df, all_fields, row_count_map


def score_priority(
    items: list[dict[str, Any]],
    all_fields: set[str],
    weights: dict[str, float] | None = None,
) -> pd.DataFrame:
    """根据字段覆盖率、问题影响、可采集性和紧迫性计算优先级。"""
    weights = weights or BASE_WEIGHTS
    rows: list[dict[str, Any]] = []
    for item in items:
        concepts = item["字段概念"]
        available_concepts: list[str] = []
        missing_concepts: list[str] = []
        for concept, aliases in concepts:
            if any(alias in all_fields for alias in aliases):
                available_concepts.append(concept)
            else:
                missing_concepts.append(concept)
        coverage_rate = len(available_concepts) / len(concepts)
        gap_rate = 1 - coverage_rate
        impact_avg = np.mean([item[f"{label}影响"] for label in QUESTION_LABELS])
        impact_norm = float(impact_avg / 5)
        feasibility_norm = float(item["可采集性评分"] / 5)
        urgency_norm = float(item["紧迫性评分"] / 5)
        score = 100 * (
            weights["影响"] * impact_norm
            + weights["缺口"] * gap_rate
            + weights["可采集性"] * feasibility_norm
            + weights["紧迫性"] * urgency_norm
        )
        rows.append(
            {
                "数据主题": item["数据主题"],
                "已有关键字段": "、".join(available_concepts) if available_concepts else "暂无",
                "建议新增字段": "、".join(missing_concepts) if missing_concepts else "暂无",
                "已有字段数": len(available_concepts),
                "需求字段数": len(concepts),
                "当前覆盖率": coverage_rate,
                "数据缺口率": gap_rate,
                "问题1影响": item["问题1影响"],
                "问题2影响": item["问题2影响"],
                "问题3影响": item["问题3影响"],
                "综合影响归一化": impact_norm,
                "可采集性评分": item["可采集性评分"],
                "紧迫性评分": item["紧迫性评分"],
                "综合优先级得分": score,
                "采集频率": item["采集频率"],
                "数据来源": item["数据来源"],
                "核心作用": item["核心作用"],
                "问题1帮助": item["问题1帮助"],
                "问题2帮助": item["问题2帮助"],
                "问题3帮助": item["问题3帮助"],
            }
        )
    result = pd.DataFrame(rows)
    # 统一保留6位小数并采用并列排名，避免浮点末位差异导致同分项目被拆到不同批次。
    result["综合优先级得分"] = result["综合优先级得分"].round(6)
    result = result.sort_values(
        ["综合优先级得分", "数据主题"],
        ascending=[False, True],
        kind="mergesort",
    ).reset_index(drop=True)
    result["综合排名"] = result["综合优先级得分"].rank(method="min", ascending=False).astype(int)
    result["实施批次"] = np.select(
        [result["综合排名"] <= 3, result["综合排名"] <= 6],
        ["第一批—立即建设", "第二批—近期建设"],
        default="第三批—持续完善",
    )
    return result


def build_impact_table(priority_df: pd.DataFrame) -> pd.DataFrame:
    impact = priority_df[
        ["数据主题", "问题1影响", "问题2影响", "问题3影响", "综合影响归一化"]
    ].copy()
    impact["综合影响得分(0-5)"] = impact["综合影响归一化"] * 5
    return impact.drop(columns=["综合影响归一化"])


def build_weight_sensitivity(all_fields: set[str]) -> pd.DataFrame:
    scenarios = {
        "基准权重": BASE_WEIGHTS,
        "影响优先": {"影响": 0.50, "缺口": 0.20, "可采集性": 0.20, "紧迫性": 0.10},
        "缺口优先": {"影响": 0.30, "缺口": 0.40, "可采集性": 0.20, "紧迫性": 0.10},
    }
    scenario_results: dict[str, pd.DataFrame] = {
        name: score_priority(DATA_REQUIREMENTS, all_fields, weights)
        for name, weights in scenarios.items()
    }
    base = scenario_results["基准权重"][["数据主题", "综合优先级得分"]].copy()
    base["基准排名"] = base["综合优先级得分"].rank(ascending=False, method="min").astype(int)
    base = base.drop(columns=["综合优先级得分"])
    for name in ["影响优先", "缺口优先"]:
        temp = scenario_results[name][["数据主题", "综合优先级得分"]].copy()
        temp[f"{name}排名"] = temp["综合优先级得分"].rank(ascending=False, method="min").astype(int)
        base = base.merge(temp.drop(columns=["综合优先级得分"]), on="数据主题", how="left")
    rank_cols = ["基准排名", "影响优先排名", "缺口优先排名"]
    base["三情景平均排名"] = base[rank_cols].mean(axis=1)
    base["前三名出现次数"] = (base[rank_cols] <= 3).sum(axis=1)
    return base.sort_values(["基准排名", "数据主题"]).reset_index(drop=True)


def run_monte_carlo(
    all_fields: set[str],
    runs: int = MONTE_CARLO_RUNS,
    seed: int = MONTE_CARLO_SEED,
) -> pd.DataFrame:
    """对权重和专家评分进行扰动，检验数据采集排序的稳定性。"""
    base_df = score_priority(DATA_REQUIREMENTS, all_fields)
    base_by_topic = base_df.set_index("数据主题")
    topics = [item["数据主题"] for item in DATA_REQUIREMENTS]
    item_count = len(topics)
    base_ordered = base_by_topic.reindex(topics)
    base_scores = base_ordered["综合优先级得分"].to_numpy(dtype=float)
    base_ranks = base_ordered["综合排名"].to_numpy(dtype=int)
    base_batches = base_ordered["实施批次"].astype(str).to_numpy()

    impact_base = np.array(
        [[item[f"{label}影响"] for label in QUESTION_LABELS] for item in DATA_REQUIREMENTS],
        dtype=float,
    )
    feasibility_base = np.array([item["可采集性评分"] for item in DATA_REQUIREMENTS], dtype=float)
    urgency_base = np.array([item["紧迫性评分"] for item in DATA_REQUIREMENTS], dtype=float)
    gap_base = base_ordered["数据缺口率"].to_numpy(dtype=float)

    rng = np.random.default_rng(seed)
    base_weight_vector = np.array(list(BASE_WEIGHTS.values()), dtype=float)
    weight_draws = base_weight_vector[None, :] * rng.uniform(
        1 - WEIGHT_PERTURBATION,
        1 + WEIGHT_PERTURBATION,
        size=(runs, len(base_weight_vector)),
    )
    weight_draws = weight_draws / weight_draws.sum(axis=1, keepdims=True)

    impact_draws = np.clip(
        impact_base[None, :, :]
        + rng.uniform(-SCORE_PERTURBATION, SCORE_PERTURBATION, size=(runs, item_count, 3)),
        0,
        5,
    )
    feasibility_draws = np.clip(
        feasibility_base[None, :]
        + rng.uniform(-SCORE_PERTURBATION, SCORE_PERTURBATION, size=(runs, item_count)),
        0,
        5,
    )
    urgency_draws = np.clip(
        urgency_base[None, :]
        + rng.uniform(-SCORE_PERTURBATION, SCORE_PERTURBATION, size=(runs, item_count)),
        0,
        5,
    )

    simulated_scores = 100 * (
        weight_draws[:, 0, None] * impact_draws.mean(axis=2) / 5
        + weight_draws[:, 1, None] * gap_base[None, :]
        + weight_draws[:, 2, None] * feasibility_draws / 5
        + weight_draws[:, 3, None] * urgency_draws / 5
    )
    order = np.argsort(-simulated_scores, axis=1, kind="stable")
    simulated_ranks = np.empty_like(simulated_scores, dtype=float)
    row_index = np.arange(runs)
    for position in range(item_count):
        simulated_ranks[row_index, order[:, position]] = position + 1

    score_quantiles = np.quantile(simulated_scores, [0.025, 0.975], axis=0)
    rank_quantiles = np.quantile(simulated_ranks, [0.025, 0.975], axis=0)
    top3_probability = (simulated_ranks <= 3).mean(axis=0)

    rows: list[dict[str, Any]] = []
    for index, topic in enumerate(topics):
        rows.append(
            {
                "数据主题": topic,
                "基准综合得分": float(base_scores[index]),
                "基准排名": int(base_ranks[index]),
                "基准实施批次": base_batches[index],
                "平均综合得分": float(simulated_scores[:, index].mean()),
                "得分标准差": float(simulated_scores[:, index].std(ddof=1)),
                "得分2.5%分位": float(score_quantiles[0, index]),
                "得分97.5%分位": float(score_quantiles[1, index]),
                "平均排名": float(simulated_ranks[:, index].mean()),
                "排名标准差": float(simulated_ranks[:, index].std(ddof=1)),
                "排名2.5%分位": float(rank_quantiles[0, index]),
                "排名97.5%分位": float(rank_quantiles[1, index]),
                "进入前三概率": float(top3_probability[index]),
                "模拟次数": int(runs),
            }
        )
    return pd.DataFrame(rows).sort_values("基准排名").reset_index(drop=True)


def build_score_notes() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "指标": "当前覆盖率",
                "定义/公式": "已有关键字段数 ÷ 需求字段总数",
                "取值范围": "0—1",
                "说明": "仅统计字段名和口径可直接对应的已有数据；日批发价和历史销量等代理字段不计入实际结算价、执行反馈销量的覆盖率。",
            },
            {
                "指标": "数据缺口率",
                "定义/公式": "1 − 当前覆盖率",
                "取值范围": "0—1",
                "说明": "缺口越大，新增数据的边际价值越高。",
            },
            {
                "指标": "评分性质",
                "定义/公式": "影响、可采集性和紧迫性均采用0—5分管理评分",
                "取值范围": "0—5",
                "说明": "评分依据前三问的建模需求、数据缺口和门店经营逻辑进行专家赋分，不是由历史样本统计估计得到的参数；覆盖率是由字段自动识别得到的客观指标。",
            },
            {
                "指标": "综合影响归一化",
                "定义/公式": "(问题1影响 + 问题2影响 + 问题3影响) ÷ 15",
                "取值范围": "0—1",
                "说明": "三道问题采用等权影响评价，避免人为偏向某一道题。",
            },
            {
                "指标": "可采集性归一化",
                "定义/公式": "可采集性评分 ÷ 5",
                "取值范围": "0—1",
                "说明": "综合考虑系统改造难度、隐私约束、采集成本和数据标准化难度。",
            },
            {
                "指标": "紧迫性归一化",
                "定义/公式": "紧迫性评分 ÷ 5",
                "取值范围": "0—1",
                "说明": "优先考虑会直接改变需求、成本、损耗或可行性判断的数据。",
            },
            {
                "指标": "综合优先级得分",
                "定义/公式": "100 × (0.40×影响 + 0.30×缺口 + 0.20×可采集性 + 0.10×紧迫性)",
                "取值范围": "0—100",
                "说明": "用于比较9类数据的相对建设价值，并作为实施批次划分依据。",
            },
            {
                "指标": "实施批次",
                "定义/公式": "采用并列排名后，排名≤3/4≤排名≤6/排名≥7",
                "取值范围": "第一批/第二批/第三批",
                "说明": "同分项目共享名次并进入同一批次，因此批次数量不强制为3、3、3；批次仅表示落地顺序。",
            },
            {
                "指标": "蒙特卡洛权重扰动",
                "定义/公式": "基准权重 × U(0.8,1.2)，再标准化",
                "取值范围": "权重和为1",
                "说明": f"共进行{MONTE_CARLO_RUNS:,}次模拟，用于检验综合优先级排序对权重设定的敏感性。",
            },
            {
                "指标": "蒙特卡洛评分扰动",
                "定义/公式": "专家评分 + U(-0.5,0.5)，并截断到[0,5]",
                "取值范围": "0—5",
                "说明": "扰动影响、可采集性和紧迫性评分；字段覆盖率由实际字段识别得到，不进行随机扰动。",
            },
            {
                "指标": "进入前三概率",
                "定义/公式": "模拟中排名≤3的次数 ÷ 模拟次数",
                "取值范围": "0—1",
                "说明": "第一批固定为排名前3名，因此该概率也表示进入第一批建设的稳定性。",
            },
        ]
    )


def save_figure(fig: plt.Figure, filename: str) -> None:
    fig.savefig(OUTPUT_DIR / filename, bbox_inches="tight", facecolor=PAPER_BG)
    plt.close(fig)


def plot_priority(priority_df: pd.DataFrame) -> None:
    data = priority_df.sort_values("综合优先级得分", ascending=True)

    fig, ax = plt.subplots(figsize=(11, 7.5), facecolor=PAPER_BG)
    ax.set_facecolor(PAPER_BG)

    bars = ax.barh(
        data["数据主题"],
        data["综合优先级得分"],
        color=[BATCH_COLORS[x] for x in data["实施批次"]],
        alpha=0.62,
        edgecolor=[BATCH_COLORS[x] for x in data["实施批次"]],
        linewidth=0.8,
    )

    ax.bar_label(
        bars,
        fmt="%.1f",
        padding=5,
        fontsize=9,
        color=TEXT_MID,
    )
    ax.set_xlim(0, 105)
    ax.set_xlabel("综合评价得分（0—100）")
    ax.set_title("第四问：数据采集综合得分与实施批次", fontsize=18, pad=16)

    ax.grid(
        axis="x",
        color=GRID_COLOR,
        alpha=0.42,
        linewidth=0.8,
        linestyle="--",
    )
    ax.set_axisbelow(True)

    # 弱化边框，使整图更轻盈
    ax.spines["top"].set_visible(False)
    ax.spines["right"].set_visible(False)
    ax.spines["left"].set_color(EDGE_COLOR)
    ax.spines["bottom"].set_color(EDGE_COLOR)

    handles = [
        patches.Patch(
            facecolor=BATCH_COLORS["第一批—立即建设"],
            edgecolor=BATCH_COLORS["第一批—立即建设"],
            alpha=0.62,
            label="第一批—立即建设",
        ),
        patches.Patch(
            facecolor=BATCH_COLORS["第二批—近期建设"],
            edgecolor=BATCH_COLORS["第二批—近期建设"],
            alpha=0.62,
            label="第二批—近期建设",
        ),
        patches.Patch(
            facecolor=BATCH_COLORS["第三批—持续完善"],
            edgecolor=BATCH_COLORS["第三批—持续完善"],
            alpha=0.62,
            label="第三批—持续完善",
        ),
    ]

    ax.legend(
        handles=handles,
        frameon=False,
        ncol=3,
        loc="lower center",
        bbox_to_anchor=(0.5, -0.16),
        labelcolor=TEXT_DARK,
    )
    fig.subplots_adjust(left=0.30, right=0.96, top=0.92, bottom=0.14)
    save_figure(fig, "图4-1_数据采集优先级.png")


def plot_impact_matrix(priority_df: pd.DataFrame) -> None:
    data = priority_df.sort_values("综合优先级得分", ascending=False)
    matrix = data[["问题1影响", "问题2影响", "问题3影响"]].to_numpy(float)

    fig, ax = plt.subplots(figsize=(8.8, 7.2), facecolor=PAPER_BG)
    ax.set_facecolor(PAPER_BG)

    image = ax.imshow(
        matrix,
        cmap=IMPACT_CMAP,
        vmin=0,
        vmax=5,
        aspect="auto",
        alpha=0.92,
    )

    ax.set_xticks(np.arange(3), QUESTION_LABELS)
    ax.set_yticks(np.arange(len(data)), data["数据主题"])
    ax.set_title("数据需求对前三问的影响矩阵", fontsize=18, pad=16)

    for i in range(matrix.shape[0]):
        for j in range(matrix.shape[1]):
            # 低饱和热力图下使用深灰文字即可，避免纯白文字突兀
            text_color = "#F7F5F1" if matrix[i, j] >= 4.5 else TEXT_DARK
            ax.text(
                j,
                i,
                f"{int(matrix[i, j])}",
                ha="center",
                va="center",
                color=text_color,
                fontsize=11,
                weight="semibold",
            )

    cbar = fig.colorbar(image, ax=ax, fraction=0.035, pad=0.04)
    cbar.set_label("影响程度（0—5分）", color=TEXT_DARK)
    cbar.ax.tick_params(colors=TEXT_MID)
    cbar.outline.set_edgecolor(EDGE_COLOR)

    ax.set_xlabel("前三问")
    ax.grid(False)

    for spine in ax.spines.values():
        spine.set_color(EDGE_COLOR)
        spine.set_alpha(0.55)

    fig.subplots_adjust(left=0.34, right=0.90, top=0.92, bottom=0.10)
    save_figure(fig, "图4-2_数据对前三问影响矩阵.png")


def plot_data_loop() -> None:
    fig, ax = plt.subplots(figsize=(13, 4.6), facecolor=PAPER_BG)
    ax.set_facecolor(PAPER_BG)
    ax.set_xlim(0, 13)
    ax.set_ylim(0, 4.2)
    ax.axis("off")

    blocks = [
        (0.35, "多源数据采集", "POS / 库存 / 采购\n损耗 / 客流 / 外部因素", FLOW_COLORS[0]),
        (3.45, "口径与质量校验", "编码匹配、缺失检查\n时间对齐、异常识别", FLOW_COLORS[1]),
        (6.55, "预测与优化决策", "需求预测、价格弹性\n补货定价与损耗控制", FLOW_COLORS[2]),
        (9.65, "执行反馈与滚动更新", "实际售价、销量、报损\n策略误差与试验结果", FLOW_COLORS[3]),
    ]

    for x, title, subtitle, color in blocks:
        rect = patches.FancyBboxPatch(
            (x, 1.25),
            2.35,
            1.75,
            boxstyle="round,pad=0.05,rounding_size=0.08",
            linewidth=1.0,
            edgecolor=EDGE_COLOR,
            facecolor=color,
            alpha=0.68,
        )
        ax.add_patch(rect)

        ax.text(
            x + 1.175,
            2.42,
            title,
            ha="center",
            va="center",
            fontsize=13,
            color=TEXT_DARK,
            weight="bold",
        )
        ax.text(
            x + 1.175,
            1.78,
            subtitle,
            ha="center",
            va="center",
            fontsize=10,
            color=TEXT_MID,
            linespacing=1.45,
        )

    for x in [2.78, 5.88, 8.98]:
        ax.annotate(
            "",
            xy=(x + 0.48, 2.12),
            xytext=(x, 2.12),
            arrowprops={
                "arrowstyle": "->",
                "lw": 1.4,
                "color": TEXT_MID,
                "alpha": 0.72,
            },
        )

    ax.annotate(
        "反馈回流",
        xy=(1.55, 1.18),
        xytext=(10.75, 0.50),
        arrowprops={
            "arrowstyle": "->",
            "lw": 1.15,
            "color": TEXT_MID,
            "alpha": 0.68,
            "connectionstyle": "arc3,rad=0.22",
        },
        ha="center",
        fontsize=11,
        color=TEXT_MID,
    )

    ax.set_title("补货与定价的数据闭环：从采集到模型更新", fontsize=18, pad=12)
    save_figure(fig, "图4-3_数据闭环与模型更新流程.png")


def plot_monte_carlo(mc_df: pd.DataFrame) -> None:
    data = mc_df.sort_values("基准排名", ascending=False).reset_index(drop=True)

    fig, (ax1, ax2) = plt.subplots(
        1,
        2,
        figsize=(15, 8),
        sharey=True,
        facecolor=PAPER_BG,
    )
    ax1.set_facecolor(PAPER_BG)
    ax2.set_facecolor(PAPER_BG)

    # 左图：进入前三概率
    bars = ax1.barh(
        data["数据主题"],
        data["进入前三概率"] * 100,
        color=MC_BAR_COLOR,
        alpha=0.60,
        edgecolor=MC_BAR_COLOR,
        linewidth=0.8,
    )
    ax1.bar_label(
        bars,
        labels=[f"{value:.1%}" for value in data["进入前三概率"]],
        padding=4,
        fontsize=9,
        color=TEXT_MID,
    )
    ax1.set_xlim(0, 105)
    ax1.set_xlabel("进入前三概率（%）")
    ax1.set_title("第一批建设建议的稳定性", fontsize=14, pad=12)
    ax1.grid(
        axis="x",
        color=GRID_COLOR,
        alpha=0.40,
        linewidth=0.8,
        linestyle="--",
    )
    ax1.set_axisbelow(True)

    # 右图：平均排名与95%区间
    y_positions = np.arange(len(data))
    rank_mean = data["平均排名"].to_numpy(float)
    rank_error = np.vstack(
        [
            rank_mean - data["排名2.5%分位"].to_numpy(float),
            data["排名97.5%分位"].to_numpy(float) - rank_mean,
        ]
    )

    ax2.errorbar(
        rank_mean,
        y_positions,
        xerr=rank_error,
        fmt="o",
        color=MC_POINT_COLOR,
        ecolor=MC_POINT_COLOR,
        elinewidth=1.5,
        capsize=3,
        markersize=5.5,
        alpha=0.72,
        markerfacecolor=MC_POINT_COLOR,
        markeredgecolor=MC_POINT_COLOR,
    )
    ax2.axvline(
        3.5,
        color=TEXT_MID,
        linestyle="--",
        linewidth=0.9,
        alpha=0.48,
    )
    ax2.set_xlim(0.5, 9.5)
    ax2.set_xticks(np.arange(1, 10))
    ax2.set_xlabel("平均排名（越靠左优先级越高，区间越窄越稳定）")
    ax2.set_title("综合排名的波动范围", fontsize=14, pad=12)
    ax2.grid(
        axis="x",
        color=GRID_COLOR,
        alpha=0.40,
        linewidth=0.8,
        linestyle="--",
    )
    ax2.set_axisbelow(True)
    ax2.tick_params(axis="y", labelleft=False)

    for ax in (ax1, ax2):
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color(EDGE_COLOR)
        ax.spines["bottom"].set_color(EDGE_COLOR)

    fig.suptitle(
        "数据采集优先级的蒙特卡洛稳健性检验",
        fontsize=18,
        y=0.97,
    )
    fig.subplots_adjust(
        left=0.30,
        right=0.97,
        top=0.88,
        bottom=0.12,
        wspace=0.42,
    )
    save_figure(fig, "图4-4_蒙特卡洛稳健性.png")


def write_model_note(
    priority_df: pd.DataFrame,
    impact_df: pd.DataFrame,
    row_count_map: dict[str, int],
    mc_df: pd.DataFrame,
) -> None:
    top_rows = [
        f"| {row['数据主题']} | {row['实施批次']} | {float(row['综合优先级得分']):.1f} | {float(row['当前覆盖率']):.1%} | {row['采集频率']} |"
        for _, row in priority_df.iterrows()
    ]
    impact_rows = [
        f"| {row['数据主题']} | {int(row['问题1影响'])} | {int(row['问题2影响'])} | {int(row['问题3影响'])} | {float(row['综合影响得分(0-5)']):.2f} |"
        for _, row in impact_df.iterrows()
    ]
    first_batch = priority_df[priority_df["实施批次"] == "第一批—立即建设"]["数据主题"].tolist()
    second_batch = priority_df[priority_df["实施批次"] == "第二批—近期建设"]["数据主题"].tolist()
    third_batch = priority_df[priority_df["实施批次"] == "第三批—持续完善"]["数据主题"].tolist()
    first_text = "、".join(first_batch) if first_batch else "无"
    second_text = "、".join(second_batch) if second_batch else "无"
    third_text = "、".join(third_batch) if third_batch else "无"
    baseline_top3 = mc_df[mc_df["基准排名"] <= 3]
    baseline_top3_probability = float(baseline_top3["进入前三概率"].mean())
    most_stable = mc_df.sort_values("进入前三概率", ascending=False).iloc[0]
    lines = [
        "# 第四问：面向补货与定价决策的数据采集优先级模型",
        "",
        "## 1. 问题理解",
        "",
        "第四问要求说明商超还应采集哪些数据，以及这些数据如何帮助解决前三问。前三问已经利用销售流水、批发价格、单品信息和近期损耗率完成了关联分析、品类预测、品类补货定价和单品补货定价，但现有数据主要记录了‘已经发生的销售’，没有完整记录缺货、库存、供应、质量和策略执行过程。因此，本问把数据采集建议转化为一个可解释的优先级评价问题。",
        "",
        "## 2. 现有数据基础与主要缺口",
        "",
        f"附件1包含 {row_count_map.get('附件1', 251):,} 个单品信息记录，附件2包含 {row_count_map.get('附件2', 878503):,} 条销售流水，附件3包含 {row_count_map.get('附件3', 55982):,} 条批发价格记录，附件4的单品损耗率表包含 251 个单品。附件2虽然有销售时间、销量、售价、销售类型和折扣标志，但没有库存和缺货状态、匿名顾客或购物篮编号；附件3只有日批发价格，没有订单、供应商和交期；附件4提供的是近期损耗率，缺少逐笔报损量和报损原因。",
        "",
        "现有字段完整并不等于决策数据完整。例如销售流水没有缺货状态时，观测销量可能只是‘实际卖出量’，而不是‘在有货条件下的潜在需求’；只有聚合损耗率时，也无法区分过量补货、品质差异和临期折价造成的损耗。",
        "",
        "## 3. 数据需求优先级模型",
        "",
        "对每类待采集数据，先根据其对问题1、问题2、问题3的作用分别给出0—5分影响评价，再根据现有字段自动计算覆盖率和缺口率。设数据项 r 的三问影响分数为 I_{r1}, I_{r2}, I_{r3}，则综合影响归一化为：",
        "",
        "$$I_r=\\frac{I_{r1}+I_{r2}+I_{r3}}{15}.$$",
        "",
        "设已有关键字段数为 A_r，所需字段总数为 B_r，则数据缺口率为：",
        "",
        "$$G_r=1-\\frac{A_r}{B_r}.$$",
        "",
        "将可采集性和紧迫性评分分别归一化到[0,1]，采用基准权重0.40、0.30、0.20、0.10，得到综合优先级：",
        "",
        "$$P_r=100(0.40I_r+0.30G_r+0.20F_r+0.10U_r).$$",
        "",
        "影响程度、可采集性和紧迫性根据前三问的建模需求、数据缺口及门店经营逻辑进行专家赋分，不属于由历史样本统计估计得到的参数；当前覆盖率则由已有数据字段自动识别得到。除三组确定性权重情景外，进一步采用蒙特卡洛模拟检验排序稳健性。",
        "",
        "## 4. 数据采集建议及作用",
        "",
        "| 数据主题 | 实施批次 | 得分 | 当前覆盖率 | 建议采集频率 |",
        "|---|---|---:|---:|---|",
        *top_rows,
        "",
        "影响矩阵如下，分数越高表示对相应问题的决策改进作用越大：",
        "",
        "| 数据主题 | 问题1 | 问题2 | 问题3 | 综合影响 |",
        "|---|---:|---:|---:|---:|",
        *impact_rows,
        "",
        f"按照基准权重，9类数据均具有较高综合建设价值。为给出可执行的实施顺序，按综合得分划分批次：第一批立即建设{first_text}；第二批近期建设{second_text}；第三批持续完善{third_text}。同分项目采用并列排名，因此批次数量不强制为3、3、3。结合本次排序，第一批重点纠正需求观测偏差并完善价格与策略执行反馈；第二批重点补充外生需求因素、真实采购约束和消费关联；第三批重点完善损耗事实、质量异质性和门店运营约束。",
        "",
        "## 5. 对前三问的具体帮助",
        "",
        "1. 对问题1：库存、缺货、促销和天气数据可以区分共同外部冲击与真正的品类关联；顾客购物篮数据可以进一步识别单品之间的替代和互补关系。",
        "2. 对问题2：促销记录、客流、天气、节假日和竞品价格可以提高未来一周需求预测和价格弹性估计；订货、到货和供应商交期可以把补货计划与真实采购过程连接起来。",
        "3. 对问题3：库存状态、质量等级、逐笔损耗和作业容量可以改进单品需求、补货量、价格和利润计算；顾客购物篮和执行反馈可以刻画未入选单品的需求转移，并验证模型效果。",
        "",
        "## 6. 数据闭环实施建议",
        "",
        "建议建立‘多源采集—口径校验—预测与优化—执行反馈’闭环。销售、库存、采购、报损和外部环境数据先按单品编码、日期和时间统一；模型输出建议售价与补货量后，回收实际执行结果、销量、报损和利润，定期更新需求参数、价格弹性和损耗参数。对于价格和促销效果，建议在合规和可控范围内保留试验组标识，以减少仅凭历史相关性进行因果判断的风险。",
        "",
        "## 7. 蒙特卡洛稳健性检验",
        "",
        f"为检验优先级结论对主观设定的敏感性，进行{MONTE_CARLO_RUNS:,}次蒙特卡洛模拟。每次将四项基准权重分别乘以U(0.8,1.2)后标准化，并将影响、可采集性和紧迫性评分分别加入U(-0.5,0.5)扰动且截断到[0,5]；字段覆盖率和数据缺口率保持实际识别结果不变。随机种子固定为{MONTE_CARLO_SEED}，因此结果可以复现。",
        "",
        f"基准排序前三名在模拟中平均进入前三的概率为{baseline_top3_probability:.1%}；进入前三概率最高的数据主题为‘{most_stable['数据主题']}’，其概率为{float(most_stable['进入前三概率']):.1%}。各数据项的得分区间、平均排名和进入前三概率详见Excel中的‘蒙特卡洛稳健性’工作表，图4-4可作为附录图使用。",
        "",
        "## 8. 局限性",
        "",
        "本问的优先级评分属于透明的管理决策评分，不等同于统计意义上的因果效应估计。影响分数、可采集性和紧迫性需要结合门店实际系统成本定期复核。匿名顾客数据应遵守隐私和最小化采集原则，外部数据也需要进行时间对齐、缺失处理和口径校验。",
        "",
        "图片文件：图4-1展示综合优先级，图4-2展示数据对前三问的影响矩阵，图4-3展示从采集到模型滚动更新的数据闭环，图4-4展示蒙特卡洛稳健性。",
    ]
    (OUTPUT_DIR / "第四问新_建模说明.md").write_text("\n".join(lines), encoding="utf-8")


def beautify_excel(path: Path) -> None:
    """沿用项目Excel风格，统一表头、字体、筛选和冻结窗格。"""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="839FB3")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row > 1 and ws.max_column > 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(name="Microsoft YaHei", size=10, color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in ws.columns:
            values = ["" if c.value is None else str(c.value) for c in col[:250]]
            max_len = max([len(value) for value in values] + [10])
            ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 10), 38)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Microsoft YaHei", size=10, color="333333")
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
        for cell in ws[1]:
            cell.fill = header_fill
        ws.row_dimensions[1].height = 30
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        # 实施批次用颜色突出，便于打开Excel后直接识别。
        if ws.title == "数据采集优先级":
            headers = {cell.value: cell.column for cell in ws[1]}
            batch_col = headers.get("实施批次")
            score_col = headers.get("综合优先级得分")
            if score_col:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, score_col).number_format = "0.0"
            if batch_col:
                batch_fills = {
                    "第一批—立即建设": "E8D2D2",
                    "第二批—近期建设": "EEE2C8",
                    "第三批—持续完善": "D9E4EC",
                }
                for row in range(2, ws.max_row + 1):
                    value = ws.cell(row, batch_col).value
                    if value in batch_fills:
                        ws.cell(row, batch_col).fill = PatternFill("solid", fgColor=batch_fills[value])
        if ws.title == "蒙特卡洛稳健性":
            headers = {cell.value: cell.column for cell in ws[1]}
            probability_col = headers.get("进入前三概率")
            if probability_col:
                for row in range(2, ws.max_row + 1):
                    ws.cell(row, probability_col).number_format = "0.0%"
    wb.save(path)


def write_excel(
    priority_df: pd.DataFrame,
    impact_df: pd.DataFrame,
    inventory_df: pd.DataFrame,
    notes_df: pd.DataFrame,
    sensitivity_df: pd.DataFrame,
    mc_df: pd.DataFrame,
) -> Path:
    path = OUTPUT_DIR / "第四问新_数据采集建议.xlsx"
    priority_columns = [
        "综合排名",
        "数据主题",
        "实施批次",
        "综合优先级得分",
        "当前覆盖率",
        "数据缺口率",
        "问题1影响",
        "问题2影响",
        "问题3影响",
        "可采集性评分",
        "紧迫性评分",
        "采集频率",
        "数据来源",
    ]
    detail_columns = [
        "数据主题",
        "已有关键字段",
        "建议新增字段",
        "需求字段数",
        "核心作用",
        "问题1帮助",
        "问题2帮助",
        "问题3帮助",
    ]
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        priority_df[priority_columns].to_excel(writer, sheet_name="数据采集优先级", index=False)
        impact_df.to_excel(writer, sheet_name="问题影响矩阵", index=False)
        inventory_df.to_excel(writer, sheet_name="现有数据盘点", index=False)
        notes_df.to_excel(writer, sheet_name="评分说明", index=False)
        sensitivity_df.to_excel(writer, sheet_name="权重敏感性", index=False)
        mc_df.to_excel(writer, sheet_name="蒙特卡洛稳健性", index=False)
        priority_df[detail_columns].to_excel(writer, sheet_name="数据作用说明", index=False)
    beautify_excel(path)
    return path


def main() -> None:
    require_inputs()
    font_name = configure_chinese_font()
    inventory_df, all_fields, row_count_map = read_current_inventory()
    priority_df = score_priority(DATA_REQUIREMENTS, all_fields)
    impact_df = build_impact_table(priority_df)
    sensitivity_df = build_weight_sensitivity(all_fields)
    mc_df = run_monte_carlo(all_fields)
    notes_df = build_score_notes()

    excel_path = write_excel(priority_df, impact_df, inventory_df, notes_df, sensitivity_df, mc_df)
    plot_priority(priority_df)
    plot_impact_matrix(priority_df)
    plot_data_loop()
    plot_monte_carlo(mc_df)
    write_model_note(priority_df, impact_df, row_count_map, mc_df)

    print("=" * 78)
    print("第四问新生成完成")
    print("=" * 78)
    print(f"论文图字体：{font_name}")
    print(f"数据主题数：{len(priority_df)}")
    batch_counts = priority_df["实施批次"].value_counts().to_dict()
    print(f"实施批次数量：{batch_counts}")
    print(f"最高得分：{priority_df.iloc[0]['数据主题']}（{float(priority_df.iloc[0]['综合优先级得分']):.2f}）")
    print(f"Excel：{excel_path}")
    print(f"输出目录：{OUTPUT_DIR}")


if __name__ == "__main__":
    main()
